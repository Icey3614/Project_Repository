# -*- coding: utf-8 -*-
"""教师/管理员视图：学生增删改查、成绩登记、统计、导入、导出、日志、回收站等。"""
import os
import time
import tkinter as tk
from tkinter import filedialog, messagebox

import ttkbootstrap as ttk

import dpi
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.figure import Figure
from openpyxl import Workbook, load_workbook

from gui_utils import dialog_header, resolve_storage, zebra
from settings_dialog import SettingsDialog
from stats import PASS_LINE, SUBJECTS as STATS_SUBJECTS, compute_statistics
from storage import mask_id_card
from student_window import ChartDialog, CompareDialog, make_person
from i18n import tr


def _log(storage, operator, role, action, target="", detail=""):
    try:
        storage.add_log(operator, role, action, target, detail)
    except Exception:
        pass


def _operator(parent):
    """从任意对话框向上查找操作人信息。"""
    node = parent
    for _ in range(4):
        teacher = getattr(node, "teacher", None)
        if teacher:
            return (
                teacher["username"],
                tr('管理员') if teacher.get("role") == "admin" else tr('教师'),
            )
        node = getattr(node, "parent", None)
    return tr('未知'), tr('未知')


class TeacherFrame(ttk.Frame):
    """教师登录后的主管理界面。"""

    COLUMNS = (
        ("student_id", "学号", 130, "center"),
        ("name", "姓名", 110, "center"),
        ("gender", "性别", 60, "center"),
        ("age", "年龄", 60, "center"),
        ("enroll_year", "入学年份", 95, "center"),
        ("id_card", "身份证号", 200, "center"),
        ("class_name", "班级", 110, "center"),
    )

    def __init__(self, parent, app, teacher):
        super().__init__(parent)
        self.app = app
        self.storage = app.storage
        self.teacher = teacher
        self.is_admin = teacher.get("role") == "admin"
        self.keyword_var = tk.StringVar()
        self._build()
        self.refresh()

    # ------------------------------------------------------------ 界面

    def _build(self):
        top = ttk.Frame(self, style="Banner.TFrame", padding=dpi.P((18, 10)))
        top.pack(fill="x")
        display_name = self.teacher["real_name"] or self.teacher["username"]
        ttk.Label(
            top,
            text=tr('欢迎，{}（{}）').format(display_name, self.teacher['username']),
            font=("Microsoft YaHei UI", 13, "bold"),
            style="Banner.TLabel",
        ).pack(side="left")
        ttk.Label(
            top, text=tr('双击学生行可快速修改'), style="BannerNote.TLabel"
        ).pack(side="right")
        role_text = tr('管理员') if self.is_admin else tr('教师')
        ttk.Label(
            top,
            text=tr('角色：{}').format(role_text),
            style="BannerNote.TLabel",
        ).pack(side="right", padx=(0, 18))

        def tb_button(parent, text, command, style=None, icon=None, side="left"):
            """工具栏按钮：统一紧凑尺寸 + 主题图标。"""
            kw = dict(
                text=text, command=command, padding=dpi.P((10, 5)), icon_size=14
            )
            if style:
                kw["style"] = style
            if icon:
                kw["icon"] = icon
            btn = ttk.Button(parent, **kw)
            btn.pack(side=side, padx=(0, 6))
            return btn

        # 第一排：搜索 + 学生管理 + 账号操作
        toolbar = ttk.Frame(self, padding=dpi.P((14, 6)))
        toolbar.pack(fill="x")
        ttk.Label(toolbar, text=tr('搜索：')).pack(side="left")
        search_entry = ttk.Entry(toolbar, textvariable=self.keyword_var, width=16)
        search_entry.pack(side="left", padx=(0, 6))
        search_entry.bind("<Return>", lambda e: self.on_search())
        tb_button(toolbar, tr('搜索'), self.on_search, style="Secondary.TButton",
                  icon="search")
        tb_button(toolbar, tr('显示全部'), self.refresh, style="Secondary.TButton",
                  icon="eye-fill")

        ttk.Separator(toolbar, orient="vertical").pack(
            side="left", fill="y", padx=10
        )

        tb_button(toolbar, tr('添加学生'), self.on_add, icon="plus-circle-fill")
        tb_button(toolbar, tr('修改学生'), self.on_edit, style="Secondary.TButton",
                  icon="pencil-square")
        tb_button(toolbar, tr('删除学生'), self.on_delete, style="Danger.TButton",
                  icon="trash-fill")

        ttk.Separator(toolbar, orient="vertical").pack(
            side="left", fill="y", padx=10
        )

        tb_button(toolbar, tr('修改密码'), self.on_change_password,
                  style="Secondary.TButton", icon="key-fill")
        if self.is_admin:
            tb_button(toolbar, tr('教师管理'), self.on_manage_teachers,
                      style="Purple.TButton", icon="person-gear")
        tb_button(toolbar, tr('退出登录'), self.app.show_login,
                  style="Secondary.TButton", icon="box-arrow-right", side="right")

        # 第二排：成绩/图表 + 统计工具 + 主题切换
        toolbar2 = ttk.Frame(self, padding=dpi.P((14, 6)))
        toolbar2.pack(fill="x")
        tb_button(toolbar2, tr('成绩登记'), self.on_scores,
                  style="Secondary.TButton", icon="clipboard-data")
        tb_button(toolbar2, tr('查看图表'), self.on_charts,
                  style="Secondary.TButton", icon="bar-chart-line-fill")
        tb_button(toolbar2, tr('成绩对比'), self.on_compare,
                  style="Green.TButton", icon="people-fill")

        ttk.Separator(toolbar2, orient="vertical").pack(
            side="left", fill="y", padx=10
        )

        tb_button(toolbar2, tr('统计分析'), self.on_statistics, icon="graph-up-arrow")
        tb_button(toolbar2, tr('批量导入'), self.on_import,
                  style="Secondary.TButton", icon="upload")
        tb_button(toolbar2, tr('导出备份'), self.on_export,
                  style="Secondary.TButton", icon="download")
        tb_button(toolbar2, tr('回收站'), self.on_recycle_bin,
                  style="Secondary.TButton", icon="recycle")
        if self.is_admin:
            tb_button(toolbar2, tr('操作日志'), self.on_logs,
                      style="Secondary.TButton", icon="journal-text")
        tb_button(
            toolbar2, tr('设置'), lambda: SettingsDialog(self, self.app),
            style="Secondary.TButton", icon="gear", side="right",
        )

        table_frame = ttk.Frame(self, padding=dpi.P((14, 6)))
        table_frame.pack(fill="both", expand=True)

        columns = [c[0] for c in self.COLUMNS]
        self.tree = ttk.Treeview(
            table_frame, columns=columns, show="headings", selectmode="extended"
        )
        for col_id, col_name, width, anchor in self.COLUMNS:
            self.tree.heading(col_id, text=tr(col_name))
            self.tree.column(
                col_id, width=dpi.scale(width), anchor=anchor, stretch=False
            )
        vsb = ttk.Scrollbar(table_frame, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(table_frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        self.tree.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")
        hsb.pack(side="bottom", fill="x")
        self.tree.bind("<Double-1>", lambda e: self.on_edit())

        self.status_var = tk.StringVar(value="")
        status_bar = ttk.Frame(self)
        status_bar.pack(side="bottom", fill="x")
        ttk.Label(
            status_bar, textvariable=self.status_var, padding=dpi.P((14, 6)),
            style="Muted.TLabel",
        ).pack(side="left")
        ttk.Button(
            status_bar, text=tr('刷新'), style="Secondary.TButton",
            icon="arrow-clockwise", command=self.refresh,
        ).pack(side="right", padx=dpi.P((0, 12)), pady=dpi.P(3))

    # ------------------------------------------------------------ 数据

    def refresh(self):
        try:
            rows = self.storage.list_students()
        except Exception as exc:
            messagebox.showerror(tr('存储错误'), str(exc))
            return
        self._fill_tree(rows, total_note=tr('共 {} 名学生').format(len(rows)))

    def on_search(self):
        keyword = self.keyword_var.get().strip()
        try:
            rows = self.storage.list_students(keyword)
        except Exception as exc:
            messagebox.showerror(tr('存储错误'), str(exc))
            return
        self._fill_tree(rows, total_note=tr('找到 {} 条匹配记录').format(len(rows)))

    def _fill_tree(self, rows, total_note=""):
        self.tree.delete(*self.tree.get_children())
        for index, row in enumerate(rows):
            self.tree.insert(
                "",
                "end",
                iid=row["student_id"],
                tags=("even" if index % 2 else "odd",),
                values=(
                    row["student_id"],
                    row["name"],
                    tr(row["gender"]),
                    "" if row["age"] is None else str(row["age"]),
                    "" if row["enroll_year"] is None else str(row["enroll_year"]),
                    mask_id_card(row.get("id_card")),
                    row.get("class_name") or "",
                ),
            )
        zebra(self.tree)
        self.status_var.set(total_note)

    # ------------------------------------------------------------ 操作

    def _selected_student(self, multiple=False):
        selected = self.tree.selection()
        if not selected:
            messagebox.showwarning(tr('提示'), tr('请先在列表中选择学生'))
            return None
        return selected if multiple else selected[0]

    def on_add(self):
        dialog = StudentDialog(self, title=tr('添加学生'))
        self.wait_window(dialog)  # 等对话框关闭后再刷新，确保新数据立即可见
        self.refresh()

    def on_edit(self):
        student_id = self._selected_student()
        if student_id is None:
            return
        try:
            student = self.storage.get_student(student_id)
        except Exception as exc:
            messagebox.showerror(tr('存储错误'), str(exc))
            return
        if student is None:
            messagebox.showwarning(tr('提示'), tr('该学生已不存在，请刷新列表'))
            self.refresh()
            return
        dialog = StudentDialog(self, title=tr('修改学生'), student=student)
        self.wait_window(dialog)
        self.refresh()

    def on_delete(self):
        selected = self._selected_student(multiple=True)
        if selected is None:
            return
        if not messagebox.askyesno(
            tr('确认删除'),
        tr('确定删除选中的 {} 名学生吗？\n将移入回收站，30 天内可恢复。').format(len(selected)),
        ):
            return
        username, role = _operator(self)
        failed = []
        for student_id in selected:
            try:
                if self.storage.delete_student(student_id):
                    _log(self.storage, username, role, tr('删除学生'), student_id)
                else:
                    failed.append(student_id)
            except Exception as exc:
                failed.append(f"{student_id}（{exc}）")
        self.refresh()
        if failed:
            messagebox.showerror(
                tr('部分删除失败'), tr('以下学生未能删除：\n') + "\n".join(failed)
            )

    def on_scores(self):
        student_id = self._selected_student()
        if student_id is None:
            return
        try:
            student = self.storage.get_student(student_id)
        except Exception as exc:
            messagebox.showerror(tr('存储错误'), str(exc))
            return
        if student is None:
            messagebox.showwarning(tr('提示'), tr('该学生已不存在，请刷新列表'))
            self.refresh()
            return
        YearlyScoresDialog(self, student)

    def on_charts(self):
        student_id = self._selected_student()
        if student_id is None:
            return
        try:
            student = self.storage.get_student(student_id)
            scores = self.storage.get_scores(student_id)
        except Exception as exc:
            messagebox.showerror(tr('存储错误'), str(exc))
            return
        if student is None:
            messagebox.showwarning(tr('提示'), tr('该学生已不存在，请刷新列表'))
            self.refresh()
            return
        if not scores:
            messagebox.showinfo(tr('提示'), tr('该学生还没有成绩记录，请先登记成绩'))
            return
        ChartDialog(self, [make_person(student, scores)])

    def on_compare(self):
        CompareDialog(self, require_self=False)

    def on_change_password(self):
        ChangePasswordDialog(self, username=self.teacher["username"])

    def on_manage_teachers(self):
        TeacherManagementDialog(self)

    def on_statistics(self):
        StatisticsDialog(self)

    def on_import(self):
        ImportDialog(self)

    def on_export(self):
        default_name = tr('学生信息备份_{}.xlsx').format(time.strftime('%Y%m%d_%H%M%S'))
        path = filedialog.asksaveasfilename(
            parent=self,
            title=tr('导出备份'),
            defaultextension=".xlsx",
            filetypes=[(tr('Excel 文件'), "*.xlsx")],
            initialfile=default_name,
        )
        if not path:
            return
        try:
            self.storage.export_to_excel(path)
        except Exception as exc:
            messagebox.showerror(tr('导出失败'), str(exc))
            return
        username, role = _operator(self)
        _log(self.storage, username, role, tr('导出备份'), os.path.basename(path))
        messagebox.showinfo(tr('成功'), tr('已导出到：\n{}').format(path))

    def on_logs(self):
        LogDialog(self)

    def on_recycle_bin(self):
        RecycleBinDialog(self)


# ==================================================================
# 学生信息添加/修改对话框
# ==================================================================

class StudentDialog(tk.Toplevel):
    def __init__(self, parent, title, student=None):
        super().__init__(parent)
        self.parent = parent
        self.storage = resolve_storage(parent)
        self.student = student
        self.title(title)
        self.resizable(False, False)
        self.transient(parent.winfo_toplevel())
        self.grab_set()
        dialog_header(self, title)
        self._build()
        self._center(parent)

    def _center(self, parent):
        self.update_idletasks()
        x = parent.winfo_rootx() + (parent.winfo_width() - self.winfo_width()) // 2
        y = parent.winfo_rooty() + (parent.winfo_height() - self.winfo_height()) // 3
        self.geometry(f"+{max(x, 0)}+{max(y, 0)}")

    def _build(self):
        is_edit = self.student is not None
        s = self.student or {}

        self.var_id = tk.StringVar(value=s.get("student_id", ""))
        self.var_name = tk.StringVar(value=s.get("name", ""))
        self.var_gender = tk.StringVar(value=s.get("gender", "男"))
        self.var_age = tk.StringVar(
            value="" if s.get("age") is None else str(s["age"])
        )
        self.var_enroll_year = tk.StringVar(
            value="" if s.get("enroll_year") is None else str(s["enroll_year"])
        )
        self.var_id_card = tk.StringVar(value=s.get("id_card") or "")
        self.var_class_name = tk.StringVar(value=s.get("class_name") or "")

        body = ttk.Frame(self, padding=dpi.P(22))
        body.pack(fill="both", expand=True)

        row = 0

        def add_entry(label, var, row_index, readonly=False):
            ttk.Label(body, text=label).grid(
                row=row_index, column=0, padx=6, pady=7, sticky="e"
            )
            ttk.Entry(
                body,
                textvariable=var,
                width=22,
                state="readonly" if readonly else "normal",
            ).grid(row=row_index, column=1, padx=6, pady=7)

        add_entry(tr('学号：'), self.var_id, row, readonly=is_edit)
        row += 1
        add_entry(tr('姓名：'), self.var_name, row)
        row += 1
        ttk.Label(body, text=tr('性别：')).grid(
            row=row, column=0, padx=6, pady=7, sticky="e"
        )
        ttk.Combobox(
            body,
            textvariable=self.var_gender,
            values=("男", "女"),
            state="readonly",
            width=20,
        ).grid(row=row, column=1, padx=6, pady=7)
        row += 1
        add_entry(tr('年龄：'), self.var_age, row)
        row += 1
        add_entry(tr('入学年份：'), self.var_enroll_year, row)
        row += 1
        add_entry(tr('身份证号：'), self.var_id_card, row)
        row += 1
        add_entry(tr('班级（可选）：'), self.var_class_name, row)
        row += 1

        ttk.Label(
            body,
            text=tr('入学年份填 4 位年份（如 2025）；身份证号 18 位'),
            style="Muted.TLabel",
        ).grid(row=row, column=0, columnspan=2, pady=(4, 10))
        row += 1

        btns = ttk.Frame(body)
        btns.grid(row=row, column=0, columnspan=2, pady=(6, 0))
        ttk.Button(btns, text=tr('保存'), width=12, command=self._save).pack(
            side="left", padx=8
        )
        ttk.Button(
            btns, text=tr('取消'), width=12, style="Secondary.TButton",
            command=self.destroy,
        ).pack(side="left", padx=8)

        self.bind("<Return>", lambda e: self._save())
        self.bind("<Escape>", lambda e: self.destroy())
        self.focus_force()

    def _save(self):
        student_id = self.var_id.get().strip()
        name = self.var_name.get().strip()
        gender = self.var_gender.get()
        age = self.var_age.get().strip()
        enroll_year = self.var_enroll_year.get().strip()
        id_card = self.var_id_card.get().strip()
        class_name = self.var_class_name.get().strip()
        username, role = _operator(self)
        try:
            if self.student is None:
                initial = self.storage.add_student(
                    student_id, name, gender, age, enroll_year, id_card,
                    class_name=class_name,
                )
                _log(self.storage, username, role, tr('添加学生'), student_id, name)
                messagebox.showinfo(
                    tr('成功'),
        tr('学生信息已保存。\n初始登录密码：{}\n（学号 + 身份证后 6 位）').format(initial),
                    parent=self,
                )
            else:
                generated = self.storage.update_student(
                    self.student["student_id"],
                    name,
                    gender,
                    age,
                    enroll_year,
                    id_card,
                    class_name,
                )
                _log(self.storage, username, role, tr('修改学生'), student_id, name)
                if generated:
                    messagebox.showinfo(
                        tr('成功'),
        tr('学生信息已保存。\n已生成初始登录密码：{}').format(generated),
                        parent=self,
                    )
                else:
                    messagebox.showinfo(tr('成功'), tr('学生信息已保存'), parent=self)
        except ValueError as exc:
            messagebox.showwarning(tr('无法保存'), str(exc), parent=self)
            return
        except Exception as exc:
            messagebox.showerror(tr('存储错误'), str(exc), parent=self)
            return
        self.destroy()


# ==================================================================
# 分年度成绩登记对话框
# ==================================================================

class YearlyScoresDialog(tk.Toplevel):
    def __init__(self, parent, student):
        super().__init__(parent)
        self.parent = parent
        self.storage = resolve_storage(parent)
        self.student = student
        self.title(tr('成绩登记 - {}（{}）').format(student['name'], student['student_id']))
        self.geometry(dpi.geom(620, 400))
        self.minsize(*dpi.minsz(560, 340))
        self.transient(parent.winfo_toplevel())
        self.grab_set()
        dialog_header(
            self,
            tr('成绩登记'),
            f"{student['name']}（{student['student_id']}）",
        )
        self._build()
        self._center(parent)
        self.reload()

    def _center(self, parent):
        self.update_idletasks()
        x = parent.winfo_rootx() + (parent.winfo_width() - self.winfo_width()) // 2
        y = parent.winfo_rooty() + (parent.winfo_height() - self.winfo_height()) // 3
        self.geometry(f"{self.winfo_width()}x{self.winfo_height()}+{max(x, 0)}+{max(y, 0)}")

    def _build(self):
        toolbar = ttk.Frame(self, padding=dpi.P((14, 4)))
        toolbar.pack(fill="x")
        ttk.Button(toolbar, text=tr('添加年份成绩'), command=self.on_add).pack(
            side="left", padx=(0, 6)
        )
        ttk.Button(
            toolbar, text=tr('修改'), style="Secondary.TButton", command=self.on_edit
        ).pack(side="left", padx=(0, 6))
        ttk.Button(
            toolbar, text=tr('删除'), style="Danger.TButton", command=self.on_delete
        ).pack(side="left")
        ttk.Button(
            toolbar, text=tr('刷新'), style="Secondary.TButton",
            icon="arrow-clockwise", command=self.reload,
        ).pack(side="left", padx=(6, 0))
        ttk.Button(
            toolbar, text=tr('关闭'), style="Secondary.TButton", command=self.destroy
        ).pack(side="right")

        table_frame = ttk.Frame(self, padding=dpi.P((14, 4)))
        table_frame.pack(fill="both", expand=True)

        columns = ("year", "chinese", "math", "english")
        self.tree = ttk.Treeview(
            table_frame, columns=columns, show="headings", selectmode="browse"
        )
        for col_id, col_name, width in (
            ("year", tr('年份'), 90),
            ("chinese", tr('语文'), 120),
            ("math", tr('数学'), 120),
            ("english", tr('英语'), 120),
        ):
            self.tree.heading(col_id, text=col_name)
            self.tree.column(col_id, width=dpi.scale(width), anchor="center")
        vsb = ttk.Scrollbar(table_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=vsb.set)
        self.tree.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")
        self.tree.bind("<Double-1>", lambda e: self.on_edit())

        self.status_var = tk.StringVar(value="")
        ttk.Label(
            self, textvariable=self.status_var, padding=dpi.P((14, 6)),
            style="Muted.TLabel",
        ).pack(side="bottom", fill="x")

    def reload(self):
        try:
            rows = self.storage.get_scores(self.student["student_id"])
        except Exception as exc:
            messagebox.showerror(tr('存储错误'), str(exc), parent=self)
            return
        self.tree.delete(*self.tree.get_children())
        for index, row in enumerate(rows):
            self.tree.insert(
                "",
                "end",
                iid=str(row["year"]),
                tags=("even" if index % 2 else "odd",),
                values=(
                    row["year"],
                    self._fmt(row["chinese_score"]),
                    self._fmt(row["math_score"]),
                    self._fmt(row["english_score"]),
                ),
            )
        zebra(self.tree)
        self.status_var.set(tr('已登记 {} 个年份的成绩（双击可修改）').format(len(rows)))

    @staticmethod
    def _fmt(value):
        if value is None:
            return ""
        text = f"{float(value):.1f}"
        return text.rstrip("0").rstrip(".")

    def on_add(self):
        dialog = ScoreDialog(self, student_id=self.student["student_id"])
        self.wait_window(dialog)  # 等成绩保存、对话框关闭后再刷新列表
        self.reload()

    def on_edit(self):
        selected = self.tree.selection()
        if not selected:
            messagebox.showwarning(tr('提示'), tr('请先选择要修改的年份'), parent=self)
            return
        year = int(selected[0])
        try:
            score = self.storage.get_score(self.student["student_id"], year)
        except Exception as exc:
            messagebox.showerror(tr('存储错误'), str(exc), parent=self)
            return
        if score is None:
            messagebox.showwarning(tr('提示'), tr('该年份记录已不存在'), parent=self)
            self.reload()
            return
        dialog = ScoreDialog(
            self, student_id=self.student["student_id"], score=score
        )
        self.wait_window(dialog)
        self.reload()

    def on_delete(self):
        selected = self.tree.selection()
        if not selected:
            messagebox.showwarning(tr('提示'), tr('请先选择要删除的年份'), parent=self)
            return
        year = int(selected[0])
        if not messagebox.askyesno(
            tr('确认删除'), tr('确定删除 {} 年的成绩记录吗？').format(year), parent=self
        ):
            return
        try:
            if not self.storage.delete_score(self.student["student_id"], year):
                messagebox.showwarning(tr('提示'), tr('该记录已不存在'), parent=self)
        except Exception as exc:
            messagebox.showerror(tr('存储错误'), str(exc), parent=self)
            return
        username, role = _operator(self)
        _log(
            self.storage, username, role, tr('删除成绩'),
            self.student["student_id"], tr('{} 年').format(year),
        )
        self.reload()


class ScoreDialog(tk.Toplevel):
    def __init__(self, parent, student_id, score=None):
        super().__init__(parent)
        self.parent = parent
        self.storage = resolve_storage(parent)
        self.student_id = student_id
        self.score = score
        self.title(tr('修改成绩') if score else tr('添加年份成绩'))
        self.resizable(False, False)
        self.transient(parent.winfo_toplevel())
        self.grab_set()
        dialog_header(self, tr('修改成绩') if score else tr('添加年份成绩'))
        self._build()
        self._center(parent)

    def _center(self, parent):
        self.update_idletasks()
        x = parent.winfo_rootx() + (parent.winfo_width() - self.winfo_width()) // 2
        y = parent.winfo_rooty() + (parent.winfo_height() - self.winfo_height()) // 3
        self.geometry(f"+{max(x, 0)}+{max(y, 0)}")

    def _build(self):
        s = self.score or {}
        self.var_year = tk.StringVar(
            value="" if s.get("year") is None else str(s["year"])
        )
        self.var_chinese = tk.StringVar(value=self._fmt(s.get("chinese_score")))
        self.var_math = tk.StringVar(value=self._fmt(s.get("math_score")))
        self.var_english = tk.StringVar(value=self._fmt(s.get("english_score")))

        body = ttk.Frame(self, padding=dpi.P(22))
        body.pack(fill="both", expand=True)

        ttk.Label(body, text=tr('年份：')).grid(
            row=0, column=0, padx=6, pady=7, sticky="e"
        )
        ttk.Entry(
            body,
            textvariable=self.var_year,
            width=22,
            state="readonly" if self.score else "normal",
        ).grid(row=0, column=1, padx=6, pady=7)

        for row, label, var in (
            (1, tr('语文成绩：'), self.var_chinese),
            (2, tr('数学成绩：'), self.var_math),
            (3, tr('英语成绩：'), self.var_english),
        ):
            ttk.Label(body, text=label).grid(
                row=row, column=0, padx=6, pady=7, sticky="e"
            )
            ttk.Entry(body, textvariable=var, width=22).grid(
                row=row, column=1, padx=6, pady=7
            )

        ttk.Label(
            body, text=tr('成绩允许留空；成绩范围 0~150'), style="Muted.TLabel"
        ).grid(row=4, column=0, columnspan=2, pady=(4, 10))

        btns = ttk.Frame(body)
        btns.grid(row=5, column=0, columnspan=2, pady=(6, 0))
        ttk.Button(btns, text=tr('保存'), width=12, command=self._save).pack(
            side="left", padx=8
        )
        ttk.Button(
            btns, text=tr('取消'), width=12, style="Secondary.TButton",
            command=self.destroy,
        ).pack(side="left", padx=8)
        self.bind("<Return>", lambda e: self._save())
        self.bind("<Escape>", lambda e: self.destroy())
        self.focus_force()

    @staticmethod
    def _fmt(value):
        return "" if value is None else str(value)

    def _save(self):
        year = self.var_year.get().strip()
        chinese = self.var_chinese.get().strip()
        math = self.var_math.get().strip()
        english = self.var_english.get().strip()
        username, role = _operator(self)
        try:
            if self.score is None:
                self.storage.add_score(self.student_id, year, chinese, math, english)
                _log(self.storage, username, role, tr('登记成绩'), self.student_id, tr('{} 年').format(year))
            else:
                self.storage.update_score(
                    self.student_id, year, chinese, math, english
                )
                _log(self.storage, username, role, tr('修改成绩'), self.student_id, tr('{} 年').format(year))
        except ValueError as exc:
            messagebox.showwarning(tr('无法保存'), str(exc), parent=self)
            return
        except Exception as exc:
            messagebox.showerror(tr('存储错误'), str(exc), parent=self)
            return
        messagebox.showinfo(tr('成功'), tr('成绩已保存'), parent=self)
        self.destroy()


# ==================================================================
# 修改教师/管理员密码对话框
# ==================================================================

class ChangePasswordDialog(tk.Toplevel):
    def __init__(self, parent, username):
        super().__init__(parent)
        self.parent = parent
        self.storage = resolve_storage(parent)
        self.username = username
        self.title(tr('修改登录密码'))
        self.resizable(False, False)
        self.transient(parent.winfo_toplevel())
        self.grab_set()
        dialog_header(self, tr('修改登录密码'), tr('账号：{}').format(username))
        self._build()
        self._center(parent)

    def _center(self, parent):
        self.update_idletasks()
        x = parent.winfo_rootx() + (parent.winfo_width() - self.winfo_width()) // 2
        y = parent.winfo_rooty() + (parent.winfo_height() - self.winfo_height()) // 3
        self.geometry(f"+{max(x, 0)}+{max(y, 0)}")

    def _build(self):
        body = ttk.Frame(self, padding=dpi.P(22))
        body.pack(fill="both", expand=True)
        ttk.Label(
            body, text=tr('正在修改账号：{}').format(self.username), style="Muted.TLabel"
        ).grid(row=0, column=0, columnspan=2, pady=(0, 12))

        self.var_old = tk.StringVar()
        self.var_new = tk.StringVar()
        self.var_confirm = tk.StringVar()

        def add_entry(label, var, row_index):
            ttk.Label(body, text=label).grid(
                row=row_index, column=0, padx=6, pady=7, sticky="e"
            )
            ttk.Entry(
                body, textvariable=var, width=22, show="*"
            ).grid(row=row_index, column=1, padx=6, pady=7)

        add_entry(tr('原密码：'), self.var_old, 1)
        add_entry(tr('新密码：'), self.var_new, 2)
        add_entry(tr('确认新密码：'), self.var_confirm, 3)
        ttk.Label(
            body, text=tr('密码长度至少 6 位，且需同时包含字母和数字'), style="Muted.TLabel"
        ).grid(row=4, column=0, columnspan=2, pady=(4, 10))

        btns = ttk.Frame(body)
        btns.grid(row=5, column=0, columnspan=2, pady=(6, 0))
        ttk.Button(btns, text=tr('确定'), width=12, command=self._save).pack(
            side="left", padx=8
        )
        ttk.Button(
            btns, text=tr('取消'), width=12, style="Secondary.TButton",
            command=self.destroy,
        ).pack(side="left", padx=8)
        self.bind("<Return>", lambda e: self._save())
        self.bind("<Escape>", lambda e: self.destroy())
        self.focus_force()

    def _save(self):
        old = self.var_old.get()
        new = self.var_new.get()
        confirm = self.var_confirm.get()
        if new != confirm:
            messagebox.showwarning(tr('提示'), tr('两次输入的新密码不一致'), parent=self)
            return
        username, role = _operator(self)
        try:
            self.storage.change_teacher_password(self.username, old, new)
        except ValueError as exc:
            messagebox.showwarning(tr('无法修改'), str(exc), parent=self)
            return
        except Exception as exc:
            messagebox.showerror(tr('存储错误'), str(exc), parent=self)
            return
        _log(self.storage, username, role, tr('修改密码'), self.username)
        messagebox.showinfo(tr('成功'), tr('密码已修改，下次登录请使用新密码'), parent=self)
        self.destroy()


# ==================================================================
# 教师账号管理（管理员专用）
# ==================================================================

ROLE_NAMES = {"teacher": "教师", "admin": "管理员"}


class TeacherManagementDialog(tk.Toplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.parent = parent
        self.app = parent.app
        self.storage = resolve_storage(parent)
        self.current_username = parent.teacher["username"]
        self.title(tr('教师账号管理'))
        self.geometry(dpi.geom(680, 430))
        self.minsize(*dpi.minsz(580, 360))
        self.transient(parent.winfo_toplevel())
        self.grab_set()
        dialog_header(self, tr('教师账号管理'))
        self._build()
        self._center(parent)
        self.reload()

    def _center(self, parent):
        self.update_idletasks()
        x = parent.winfo_rootx() + (parent.winfo_width() - self.winfo_width()) // 2
        y = parent.winfo_rooty() + (parent.winfo_height() - self.winfo_height()) // 3
        self.geometry(f"{self.winfo_width()}x{self.winfo_height()}+{max(x, 0)}+{max(y, 0)}")

    def _build(self):
        toolbar = ttk.Frame(self, padding=dpi.P((14, 10)))
        toolbar.pack(fill="x")
        ttk.Button(toolbar, text=tr('添加账号'), command=self.on_add).pack(
            side="left", padx=(0, 6)
        )
        ttk.Button(
            toolbar, text=tr('修改'), style="Secondary.TButton", command=self.on_edit
        ).pack(side="left", padx=(0, 6))
        ttk.Button(
            toolbar, text=tr('删除'), style="Danger.TButton", command=self.on_delete
        ).pack(side="left")
        ttk.Button(
            toolbar, text=tr('刷新'), style="Secondary.TButton",
            icon="arrow-clockwise", command=self.reload,
        ).pack(side="left", padx=(6, 0))
        ttk.Button(
            toolbar, text=tr('关闭'), style="Secondary.TButton", command=self.destroy
        ).pack(side="right")

        table_frame = ttk.Frame(self, padding=dpi.P((14, 4)))
        table_frame.pack(fill="both", expand=True)
        columns = ("username", "real_name", "employee_id", "role")
        self.tree = ttk.Treeview(
            table_frame, columns=columns, show="headings", selectmode="browse"
        )
        for col_id, col_name, width in (
            ("username", tr('账号'), 130),
            ("real_name", tr('姓名'), 110),
            ("employee_id", tr('工号'), 100),
            ("role", tr('角色'), 90),
        ):
            self.tree.heading(col_id, text=col_name)
            self.tree.column(col_id, width=dpi.scale(width), anchor="center")
        vsb = ttk.Scrollbar(table_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=vsb.set)
        self.tree.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")
        self.tree.bind("<Double-1>", lambda e: self.on_edit())

        self.status_var = tk.StringVar(value="")
        ttk.Label(
            self, textvariable=self.status_var, padding=dpi.P((14, 6)),
            style="Muted.TLabel",
        ).pack(side="bottom", fill="x")

    def reload(self):
        try:
            rows = self.storage.list_teachers()
        except Exception as exc:
            messagebox.showerror(tr('存储错误'), str(exc), parent=self)
            return
        self.tree.delete(*self.tree.get_children())
        for index, row in enumerate(rows):
            self.tree.insert(
                "",
                "end",
                iid=row["username"],
                tags=("even" if index % 2 else "odd",),
                values=(
                    row["username"],
                    row.get("real_name") or "—",
                    row.get("employee_id") or "—",
                    tr(ROLE_NAMES.get(row.get("role"), "教师")),
                ),
            )
        zebra(self.tree)
        self.status_var.set(tr('共 {} 个账号').format(len(rows)))

    def on_add(self):
        dialog = AddTeacherDialog(self)
        self.wait_window(dialog)  # 等教师账号创建、对话框关闭后再刷新
        self.reload()

    def on_edit(self):
        selected = self.tree.selection()
        if not selected:
            messagebox.showwarning(tr('提示'), tr('请先选择要修改的账号'), parent=self)
            return
        username = selected[0]
        row = next(
            (r for r in self.storage.list_teachers() if r["username"] == username),
            None,
        )
        if row is None:
            messagebox.showwarning(tr('提示'), tr('该账号已不存在'), parent=self)
            self.reload()
            return
        dialog = TeacherEditDialog(self, row)
        self.wait_window(dialog)
        self.reload()

    def on_delete(self):
        selected = self.tree.selection()
        if not selected:
            messagebox.showwarning(tr('提示'), tr('请先选择要删除的账号'), parent=self)
            return
        username = selected[0]
        if username == self.current_username:
            messagebox.showwarning(tr('提示'), tr('不能删除当前登录的账号'), parent=self)
            return
        if not messagebox.askyesno(
            tr('确认删除'), tr('确定删除账号 {} 吗？删除后不可恢复。').format(username), parent=self
        ):
            return
        try:
            self.storage.delete_teacher(username)
        except Exception as exc:
            messagebox.showerror(tr('存储错误'), str(exc), parent=self)
            return
        op, role = _operator(self)
        _log(self.storage, op, role, tr('删除教师'), username)
        self.reload()


class AddTeacherDialog(tk.Toplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.parent = parent
        self.storage = resolve_storage(parent)
        self.title(tr('添加教师账号'))
        self.resizable(False, False)
        self.transient(parent.winfo_toplevel())
        self.grab_set()
        dialog_header(self, tr('添加教师账号'))
        self._build()
        self._center(parent)

    def _center(self, parent):
        self.update_idletasks()
        x = parent.winfo_rootx() + (parent.winfo_width() - self.winfo_width()) // 2
        y = parent.winfo_rooty() + (parent.winfo_height() - self.winfo_height()) // 3
        self.geometry(f"+{max(x, 0)}+{max(y, 0)}")

    def _build(self):
        body = ttk.Frame(self, padding=dpi.P(22))
        body.pack(fill="both", expand=True)

        self.var_username = tk.StringVar()
        self.var_realname = tk.StringVar()
        self.var_employee = tk.StringVar()
        self.var_id_card = tk.StringVar()
        self.var_role = tk.StringVar(value=tr('教师'))

        ttk.Label(body, text=tr('登录账号：')).grid(
            row=0, column=0, padx=6, pady=7, sticky="e"
        )
        ttk.Entry(body, textvariable=self.var_username, width=22).grid(
            row=0, column=1, padx=6, pady=7
        )
        ttk.Label(body, text=tr('姓名（可选）：')).grid(
            row=1, column=0, padx=6, pady=7, sticky="e"
        )
        ttk.Entry(body, textvariable=self.var_realname, width=22).grid(
            row=1, column=1, padx=6, pady=7
        )
        ttk.Label(body, text=tr('工号：')).grid(
            row=2, column=0, padx=6, pady=7, sticky="e"
        )
        ttk.Entry(body, textvariable=self.var_employee, width=22).grid(
            row=2, column=1, padx=6, pady=7
        )
        ttk.Label(body, text=tr('身份证号：')).grid(
            row=3, column=0, padx=6, pady=7, sticky="e"
        )
        ttk.Entry(body, textvariable=self.var_id_card, width=22).grid(
            row=3, column=1, padx=6, pady=7
        )
        ttk.Label(body, text=tr('角色：')).grid(
            row=4, column=0, padx=6, pady=7, sticky="e"
        )
        ttk.Combobox(
            body,
            textvariable=self.var_role,
            values=(tr('教师'), tr('管理员')),
            state="readonly",
            width=20,
        ).grid(row=4, column=1, padx=6, pady=7)
        ttk.Label(
            body,
            text=tr('初始登录密码 = 工号 + 身份证后 6 位，创建后请转告本人'),
            style="Muted.TLabel",
        ).grid(row=5, column=0, columnspan=2, pady=(4, 10))

        btns = ttk.Frame(body)
        btns.grid(row=6, column=0, columnspan=2, pady=(6, 0))
        ttk.Button(btns, text=tr('确定'), width=12, command=self._save).pack(
            side="left", padx=8
        )
        ttk.Button(
            btns, text=tr('取消'), width=12, style="Secondary.TButton",
            command=self.destroy,
        ).pack(side="left", padx=8)
        self.bind("<Return>", lambda e: self._save())
        self.bind("<Escape>", lambda e: self.destroy())
        self.focus_force()

    def _save(self):
        username = self.var_username.get().strip()
        real_name = self.var_realname.get().strip()
        employee_id = self.var_employee.get().strip()
        id_card = self.var_id_card.get().strip()
        role = "admin" if self.var_role.get() == tr('管理员') else "teacher"
        try:
            initial = self.storage.add_teacher(
                username, real_name, None, role, employee_id, id_card
            )
        except ValueError as exc:
            messagebox.showwarning(tr('无法添加'), str(exc), parent=self)
            return
        except Exception as exc:
            messagebox.showerror(tr('存储错误'), str(exc), parent=self)
            return
        op, op_role = _operator(self)
        _log(self.storage, op, op_role, tr('添加教师'), username)
        messagebox.showinfo(
            tr('成功'),
        tr('教师账号 {} 已创建。\n初始登录密码：{}\n（工号 + 身份证后 6 位）').format(username, initial),
            parent=self,
        )
        self.destroy()


class TeacherEditDialog(tk.Toplevel):
    def __init__(self, parent, account):
        super().__init__(parent)
        self.parent = parent
        self.storage = resolve_storage(parent)
        self.account = account
        self.title(tr('修改账号 - {}').format(account['username']))
        self.resizable(False, False)
        self.transient(parent.winfo_toplevel())
        self.grab_set()
        dialog_header(self, tr('修改账号 - {}').format(account['username']))
        self._build()
        self._center(parent)

    def _center(self, parent):
        self.update_idletasks()
        x = parent.winfo_rootx() + (parent.winfo_width() - self.winfo_width()) // 2
        y = parent.winfo_rooty() + (parent.winfo_height() - self.winfo_height()) // 3
        self.geometry(f"+{max(x, 0)}+{max(y, 0)}")

    def _build(self):
        body = ttk.Frame(self, padding=dpi.P(22))
        body.pack(fill="both", expand=True)

        self.var_username = tk.StringVar(value=self.account["username"])
        self.var_realname = tk.StringVar(value=self.account.get("real_name") or "")
        self.var_employee = tk.StringVar(value=self.account.get("employee_id") or "")
        self.var_id_card = tk.StringVar(value=self.account.get("id_card") or "")
        self.var_role = tk.StringVar(
            value=tr(ROLE_NAMES.get(self.account.get("role"), "教师"))
        )
        self.var_password = tk.StringVar()
        self.var_confirm = tk.StringVar()

        ttk.Label(body, text=tr('账号：')).grid(
            row=0, column=0, padx=6, pady=7, sticky="e"
        )
        ttk.Entry(
            body, textvariable=self.var_username, width=22, state="readonly"
        ).grid(row=0, column=1, padx=6, pady=7)
        ttk.Label(body, text=tr('姓名（可选）：')).grid(
            row=1, column=0, padx=6, pady=7, sticky="e"
        )
        ttk.Entry(body, textvariable=self.var_realname, width=22).grid(
            row=1, column=1, padx=6, pady=7
        )
        ttk.Label(body, text=tr('工号：')).grid(
            row=2, column=0, padx=6, pady=7, sticky="e"
        )
        ttk.Entry(body, textvariable=self.var_employee, width=22).grid(
            row=2, column=1, padx=6, pady=7
        )
        ttk.Label(body, text=tr('身份证号：')).grid(
            row=3, column=0, padx=6, pady=7, sticky="e"
        )
        ttk.Entry(body, textvariable=self.var_id_card, width=22).grid(
            row=3, column=1, padx=6, pady=7
        )
        ttk.Label(body, text=tr('角色：')).grid(
            row=4, column=0, padx=6, pady=7, sticky="e"
        )
        ttk.Combobox(
            body,
            textvariable=self.var_role,
            values=(tr('教师'), tr('管理员')),
            state="readonly",
            width=20,
        ).grid(row=4, column=1, padx=6, pady=7)
        ttk.Label(body, text=tr('新密码（留空不改）：')).grid(
            row=5, column=0, padx=6, pady=7, sticky="e"
        )
        ttk.Entry(body, textvariable=self.var_password, width=22, show="*").grid(
            row=5, column=1, padx=6, pady=7
        )
        ttk.Label(body, text=tr('确认新密码：')).grid(
            row=6, column=0, padx=6, pady=7, sticky="e"
        )
        ttk.Entry(body, textvariable=self.var_confirm, width=22, show="*").grid(
            row=6, column=1, padx=6, pady=7
        )

        btns = ttk.Frame(body)
        btns.grid(row=7, column=0, columnspan=2, pady=(12, 0))
        ttk.Button(btns, text=tr('保存'), width=12, command=self._save).pack(
            side="left", padx=8
        )
        ttk.Button(
            btns, text=tr('取消'), width=12, style="Secondary.TButton",
            command=self.destroy,
        ).pack(side="left", padx=8)
        self.bind("<Return>", lambda e: self._save())
        self.bind("<Escape>", lambda e: self.destroy())
        self.focus_force()

    def _save(self):
        real_name = self.var_realname.get().strip()
        employee_id = self.var_employee.get().strip()
        id_card = self.var_id_card.get().strip()
        role = "admin" if self.var_role.get() == tr('管理员') else "teacher"
        new_password = self.var_password.get()
        confirm = self.var_confirm.get()
        if new_password != confirm:
            messagebox.showwarning(tr('提示'), tr('两次输入的新密码不一致'), parent=self)
            return
        try:
            self.storage.update_teacher(
                self.account["username"],
                real_name,
                role,
                new_password or None,
                employee_id,
                id_card,
            )
        except ValueError as exc:
            messagebox.showwarning(tr('无法保存'), str(exc), parent=self)
            return
        except Exception as exc:
            messagebox.showerror(tr('存储错误'), str(exc), parent=self)
            return
        op, op_role = _operator(self)
        _log(self.storage, op, op_role, tr('修改教师'), self.account["username"])
        messagebox.showinfo(tr('成功'), tr('账号信息已保存'), parent=self)
        self.destroy()


# ==================================================================
# 统计分析对话框
# ==================================================================

class StatisticsDialog(tk.Toplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.parent = parent
        self.storage = resolve_storage(parent)
        self.title(tr('成绩统计分析'))
        self.geometry(dpi.geom(1000, 700))
        self.minsize(*dpi.minsz(880, 620))
        self.transient(parent.winfo_toplevel())
        self.grab_set()
        dialog_header(self, tr('成绩统计分析'))
        self._load_data()
        self._build()
        self._center(parent)
        self._run()

    def _center(self, parent):
        self.update_idletasks()
        x = parent.winfo_rootx() + (parent.winfo_width() - self.winfo_width()) // 2
        y = parent.winfo_rooty() + (parent.winfo_height() - self.winfo_height()) // 3
        self.geometry(f"{self.winfo_width()}x{self.winfo_height()}+{max(x, 0)}+{max(y, 0)}")

    def _load_data(self):
        self.students = self.storage.list_students()
        self.all_scores = self.storage.get_all_scores()
        self.years = sorted({r["year"] for r in self.all_scores})
        classes = {
            (s.get("class_name") or "").strip()
            for s in self.students
            if (s.get("class_name") or "").strip()
        }
        self.classes = sorted(classes)

    def _build(self):
        body = ttk.Frame(self, padding=12)
        body.pack(fill="both", expand=True)

        config = ttk.Frame(body)
        config.pack(fill="x", pady=(0, 8))
        ttk.Label(config, text=tr('学年：')).pack(side="left")
        self.year_var = tk.StringVar(
            value=str(self.years[-1]) if self.years else ""
        )
        ttk.Combobox(
            config,
            textvariable=self.year_var,
            values=[str(y) for y in self.years],
            state="readonly",
            width=10,
        ).pack(side="left", padx=(0, 14))
        ttk.Label(config, text=tr('班级：')).pack(side="left")
        self.class_var = tk.StringVar(value=tr('全部班级'))
        ttk.Combobox(
            config,
            textvariable=self.class_var,
            values=[tr('全部班级')] + self.classes,
            state="readonly",
            width=14,
        ).pack(side="left", padx=(0, 14))
        ttk.Button(config, text=tr('统计'), command=self._run).pack(side="left")
        ttk.Label(
            config,
            text=tr('及格线：{} 分').format(PASS_LINE),
            style="Muted.TLabel",
        ).pack(side="right")

        self.summary_var = tk.StringVar(value="")
        ttk.Label(
            body, textvariable=self.summary_var, wraplength=960
        ).pack(fill="x", pady=(0, 8))

        table_frame = ttk.Frame(body)
        table_frame.pack(fill="both", expand=True)
        columns = ("rank", "sid", "name", "cls", "chinese", "math", "english", "total", "avg")
        self.tree = ttk.Treeview(
            table_frame, columns=columns, show="headings", selectmode="browse"
        )
        for col_id, col_name, width in (
            ("rank", tr('排名'), 60),
            ("sid", tr('学号'), 110),
            ("name", tr('姓名'), 100),
            ("cls", tr('班级'), 90),
            ("chinese", tr('语文'), 80),
            ("math", tr('数学'), 80),
            ("english", tr('英语'), 80),
            ("total", tr('总分'), 80),
            ("avg", tr('平均分'), 80),
        ):
            self.tree.heading(col_id, text=col_name)
            self.tree.column(col_id, width=dpi.scale(width), anchor="center")
        vsb = ttk.Scrollbar(table_frame, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(table_frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        self.tree.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")
        hsb.pack(side="bottom", fill="x")

        self.chart_frame = ttk.Frame(body)
        self.chart_frame.pack(fill="x", pady=(8, 0))

    def _run(self):
        if not self.years:
            self.summary_var.set(tr('暂无成绩数据'))
            return
        year = int(self.year_var.get())
        class_name = self.class_var.get()
        if class_name == tr('全部班级'):
            class_name = None
        result = compute_statistics(
            self.students, self.all_scores, year, class_name
        )

        self.tree.delete(*self.tree.get_children())
        for index, item in enumerate(result["rows"]):
            s = item["student"]
            self.tree.insert(
                "",
                "end",
                tags=("even" if index % 2 else "odd",),
                values=(
                    item["rank"],
                    s["student_id"],
                    s["name"],
                    s.get("class_name") or "",
                    self._fmt(item["语文"]),
                    self._fmt(item["数学"]),
                    self._fmt(item["英语"]),
                    f"{item['total']:.1f}",
                    f"{item['avg']:.1f}",
                ),
            )
        zebra(self.tree)

        lines = [tr('共 {} 人').format(result['count'])]
        for label in STATS_SUBJECTS:
            info = result["summary"][label]
            if info:
                lines.append(
                    tr('{}：平均 {} / 最高 {} / 最低 {} / 及格率 {}%').format(
                        tr(label), info['avg'], info['max'], info['min'],
                        info['pass_rate'],
                    )
                )
            else:
                lines.append(tr('{}：无数据').format(tr(label)))
        self.summary_var.set("　".join(lines))
        self._draw_chart(year, class_name, result)

    def _draw_chart(self, year, class_name, result):
        for child in self.chart_frame.winfo_children():
            child.destroy()
        fig = Figure(figsize=(9.5, 1.9), dpi=100)
        ax = fig.add_subplot(111)
        subject_keys = list(STATS_SUBJECTS.keys())
        labels = [tr(k) for k in subject_keys]
        values = [
            result["summary"][k]["avg"] if result["summary"][k] else 0
            for k in subject_keys
        ]
        ax.bar(labels, values, color=["#2563eb", "#16a34a", "#dc2626"], width=0.45)
        ax.set_ylim(0, 160)
        ax.grid(axis="y", linestyle="--", alpha=0.4)
        ax.set_title(
            tr('{} 年{} 三科平均分').format(year, '（' + class_name + '）' if class_name else ''),
            fontsize=11,
        )
        for i, v in enumerate(values):
            if v:
                ax.annotate(
                    f"{v:.1f}",
                    xy=(i, v),
                    xytext=(0, 3),
                    textcoords="offset points",
                    ha="center",
                    fontsize=9,
                )
        canvas = FigureCanvasTkAgg(fig, master=self.chart_frame)
        canvas.draw()
        canvas.get_tk_widget().pack(fill="both", expand=True)

    @staticmethod
    def _fmt(value):
        if value is None:
            return "—"
        return f"{float(value):.1f}".rstrip("0").rstrip(".")


# ==================================================================
# 批量导入对话框
# ==================================================================

IMPORT_HEADERS = [tr('学号'), tr('姓名'), tr('性别'), tr('年龄'), tr('入学年份'), tr('身份证号'), tr('班级')]
REQUIRED_IMPORT = [tr('学号'), tr('姓名'), tr('性别'), tr('入学年份'), tr('身份证号')]


class ImportDialog(tk.Toplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.parent = parent
        self.storage = resolve_storage(parent)
        self.rows = []
        self.title(tr('批量导入学生'))
        self.geometry(dpi.geom(600, 360))
        self.resizable(False, False)
        self.transient(parent.winfo_toplevel())
        self.grab_set()
        dialog_header(self, tr('批量导入学生'))
        self._build()
        self._center(parent)

    def _center(self, parent):
        self.update_idletasks()
        x = parent.winfo_rootx() + (parent.winfo_width() - self.winfo_width()) // 2
        y = parent.winfo_rooty() + (parent.winfo_height() - self.winfo_height()) // 3
        self.geometry(f"+{max(x, 0)}+{max(y, 0)}")

    def _build(self):
        body = ttk.Frame(self, padding=20)
        body.pack(fill="both", expand=True)

        ttk.Label(
            body,
            text=tr('选择包含学生信息的 Excel 文件（需含表头：学号、姓名、性别、入学年份、身份证号；年龄、班级可选）。'),
            wraplength=540,
            style="Muted.TLabel",
        ).pack(anchor="w", pady=(0, 12))

        self.var_path = tk.StringVar()
        row = ttk.Frame(body)
        row.pack(fill="x")
        ttk.Entry(row, textvariable=self.var_path, width=44).pack(side="left")
        ttk.Button(
            row, text=tr('浏览…'), style="Secondary.TButton", command=self._browse
        ).pack(side="left", padx=(8, 0))

        self.info_var = tk.StringVar(value=tr('尚未选择文件'))
        ttk.Label(body, textvariable=self.info_var, style="Muted.TLabel").pack(
            anchor="w", pady=(12, 8)
        )

        btns = ttk.Frame(body)
        btns.pack(fill="x", pady=(14, 0))
        ttk.Button(btns, text=tr('生成导入模板'), command=self._template).pack(
            side="left", padx=(0, 8)
        )
        ttk.Button(btns, text=tr('开始导入'), command=self._import).pack(
            side="left", padx=(0, 8)
        )
        ttk.Button(
            btns, text=tr('关闭'), style="Secondary.TButton", command=self.destroy
        ).pack(side="right")

    def _browse(self):
        path = filedialog.askopenfilename(
            parent=self,
            title=tr('选择 Excel 文件'),
            filetypes=[(tr('Excel 文件'), "*.xlsx *.xlsm")],
        )
        if path:
            self.var_path.set(path)
            try:
                self.rows = self._parse_file(path)
            except ValueError as exc:
                self.rows = []
                self.info_var.set(str(exc))
                return
            self.info_var.set(
                tr('读取到 {} 条学生记录，点击「开始导入」写入系统').format(len(self.rows))
            )

    def _parse_file(self, path):
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        header = [
            str(c).strip() if c is not None else ""
            for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        ]
        index = {
            name: header.index(name)
            for name in IMPORT_HEADERS
            if name in header
        }
        missing = [h for h in REQUIRED_IMPORT if h not in index]
        if missing:
            wb.close()
            raise ValueError(tr('文件缺少必需列：{}').format('、'.join(missing)))

        rows = []
        for raw in ws.iter_rows(min_row=2, values_only=True):
            if not raw or raw[index["学号"]] in (None, ""):
                continue

            def get(name):
                i = index.get(name)
                if i is None or i >= len(raw) or raw[i] is None:
                    return ""
                return str(raw[i]).strip()

            rows.append(
                {
                    "student_id": get("学号"),
                    "name": get("姓名"),
                    "gender": get("性别"),
                    "age": get("年龄"),
                    "enroll_year": get("入学年份"),
                    "id_card": get("身份证号"),
                    "class_name": get("班级"),
                }
            )
        wb.close()
        return rows

    def _template(self):
        path = filedialog.asksaveasfilename(
            parent=self,
            title=tr('保存导入模板'),
            defaultextension=".xlsx",
            filetypes=[(tr('Excel 文件'), "*.xlsx")],
            initialfile=tr('学生导入模板.xlsx'),
        )
        if not path:
            return
        wb = Workbook()
        ws = wb.active
        ws.title = tr('学生导入模板')
        ws.append(IMPORT_HEADERS)
        ws.append(
            ["20260101", "张三", "男", 16, 2025, "110101200801011234", "高一(1)班"]
        )
        wb.save(path)
        wb.close()
        messagebox.showinfo(tr('成功'), tr('模板已保存：\n{}').format(path), parent=self)

    def _import(self):
        if not self.rows:
            messagebox.showwarning(tr('提示'), tr('请先选择并读取一个 Excel 文件'), parent=self)
            return
        success = 0
        failures = []
        for row in self.rows:
            try:
                self.storage.add_student(
                    row["student_id"],
                    row["name"],
                    row["gender"] or "男",
                    row["age"],
                    row["enroll_year"],
                    row["id_card"],
                    class_name=row["class_name"],
                )
                success += 1
            except ValueError as exc:
                failures.append(f"{row['student_id']} {row['name']}：{exc}")
            except Exception as exc:
                failures.append(f"{row['student_id']} {row['name']}：{exc}")
        op, role = _operator(self)
        _log(
            self.storage, op, role, tr('批量导入'),
            "", tr('成功 {} 条，失败 {} 条').format(success, len(failures)),
        )
        if failures:
            messagebox.showwarning(
                tr('导入完成（部分失败）'),
        tr('成功 {} 条，失败 {} 条：\n').format(success, len(failures))
                + "\n".join(failures[:15])
                + ("\n……" if len(failures) > 15 else ""),
                parent=self,
            )
        else:
            messagebox.showinfo(tr('导入完成'), tr('成功导入 {} 名学生').format(success), parent=self)


# ==================================================================
# 操作日志对话框（管理员）
# ==================================================================

class LogDialog(tk.Toplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.parent = parent
        self.storage = resolve_storage(parent)
        self.is_admin = bool(getattr(parent, "is_admin", False))
        self.title(tr('操作日志'))
        self.geometry(dpi.geom(900, 520))
        self.minsize(*dpi.minsz(760, 420))
        self.transient(parent.winfo_toplevel())
        self.grab_set()
        dialog_header(self, tr('操作日志'))
        self._build()
        self._center(parent)
        self.reload()

    def _center(self, parent):
        self.update_idletasks()
        x = parent.winfo_rootx() + (parent.winfo_width() - self.winfo_width()) // 2
        y = parent.winfo_rooty() + (parent.winfo_height() - self.winfo_height()) // 3
        self.geometry(f"{self.winfo_width()}x{self.winfo_height()}+{max(x, 0)}+{max(y, 0)}")

    def _build(self):
        toolbar = ttk.Frame(self, padding=dpi.P((14, 10)))
        toolbar.pack(fill="x")
        ttk.Button(
            toolbar, text=tr('刷新'), style="Secondary.TButton", command=self.reload
        ).pack(side="left")
        if self.is_admin:
            ttk.Button(
                toolbar, text=tr('清空'), style="Danger.TButton",
                icon="trash-fill", command=self.on_clear,
            ).pack(side="left", padx=(6, 0))
        ttk.Button(
            toolbar, text=tr('关闭'), style="Secondary.TButton", command=self.destroy
        ).pack(side="right")

        table_frame = ttk.Frame(self, padding=dpi.P((14, 4)))
        table_frame.pack(fill="both", expand=True)
        columns = ("time", "operator", "role", "action", "target", "detail")
        self.tree = ttk.Treeview(
            table_frame, columns=columns, show="headings", selectmode="browse"
        )
        for col_id, col_name, width in (
            ("time", tr('时间'), 160),
            ("operator", tr('操作人'), 120),
            ("role", tr('角色'), 80),
            ("action", tr('操作'), 100),
            ("target", tr('对象'), 130),
            ("detail", tr('详情'), 220),
        ):
            self.tree.heading(col_id, text=col_name)
            self.tree.column(col_id, width=dpi.scale(width), anchor="w")
        vsb = ttk.Scrollbar(table_frame, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(table_frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        self.tree.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")
        hsb.pack(side="bottom", fill="x")

        self.status_var = tk.StringVar(value="")
        ttk.Label(
            self, textvariable=self.status_var, padding=dpi.P((14, 6)),
            style="Muted.TLabel",
        ).pack(side="bottom", fill="x")

    def reload(self):
        try:
            rows = self.storage.list_logs(500)
        except Exception as exc:
            messagebox.showerror(tr('存储错误'), str(exc), parent=self)
            return
        self.tree.delete(*self.tree.get_children())
        for index, row in enumerate(rows):
            self.tree.insert(
                "",
                "end",
                tags=("even" if index % 2 else "odd",),
                values=(
                    str(row.get("created_at"))[:19],
                    row.get("operator", ""),
                    row.get("role", ""),
                    row.get("action", ""),
                    row.get("target", ""),
                    row.get("detail", ""),
                ),
            )
        zebra(self.tree)
        self.status_var.set(tr('共 {} 条日志（最多显示 500 条）').format(len(rows)))

    def on_clear(self):
        """清空全部操作日志（仅管理员可操作）。"""
        if not self.is_admin:
            messagebox.showerror(tr('无权限'), tr('只有管理员可以清空操作日志'), parent=self)
            return
        if not messagebox.askyesno(
            tr('确认清空'),
            tr('确定清空全部操作日志吗？此操作不可恢复。'),
            parent=self,
        ):
            return
        try:
            self.storage.clear_logs()
        except Exception as exc:
            messagebox.showerror(tr('存储错误'), str(exc), parent=self)
            return
        # 清空后保留一条审计记录
        username, role = _operator(self)
        _log(self.storage, username, role, tr('清空日志'), "", tr('清空全部操作日志'))
        messagebox.showinfo(tr('成功'), tr('操作日志已清空'), parent=self)
        self.reload()


# ==================================================================
# 回收站对话框
# ==================================================================

class RecycleBinDialog(tk.Toplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.parent = parent
        self.storage = resolve_storage(parent)
        self.title(tr('回收站'))
        self.geometry(dpi.geom(720, 440))
        self.minsize(*dpi.minsz(620, 360))
        self.transient(parent.winfo_toplevel())
        self.grab_set()
        dialog_header(self, tr('回收站'))
        self._build()
        self._center(parent)
        self.reload()

    def _center(self, parent):
        self.update_idletasks()
        x = parent.winfo_rootx() + (parent.winfo_width() - self.winfo_width()) // 2
        y = parent.winfo_rooty() + (parent.winfo_height() - self.winfo_height()) // 3
        self.geometry(f"{self.winfo_width()}x{self.winfo_height()}+{max(x, 0)}+{max(y, 0)}")

    def _build(self):
        toolbar = ttk.Frame(self, padding=dpi.P((14, 10)))
        toolbar.pack(fill="x")
        ttk.Button(toolbar, text=tr('恢复'), command=self.on_restore).pack(
            side="left", padx=(0, 6)
        )
        ttk.Button(
            toolbar, text=tr('彻底删除'), style="Danger.TButton", command=self.on_purge
        ).pack(side="left", padx=(0, 6))
        ttk.Button(
            toolbar, text=tr('刷新'), style="Secondary.TButton", command=self.reload
        ).pack(side="left")
        ttk.Button(
            toolbar, text=tr('关闭'), style="Secondary.TButton", command=self.destroy
        ).pack(side="right")
        ttk.Label(
            toolbar,
            text=tr('超过 30 天的记录会自动彻底清除'),
            style="Muted.TLabel",
        ).pack(side="right", padx=(0, 14))

        table_frame = ttk.Frame(self, padding=dpi.P((14, 4)))
        table_frame.pack(fill="both", expand=True)
        columns = ("sid", "name", "gender", "cls", "deleted_at")
        self.tree = ttk.Treeview(
            table_frame, columns=columns, show="headings", selectmode="extended"
        )
        for col_id, col_name, width in (
            ("sid", tr('学号'), 130),
            ("name", tr('姓名'), 110),
            ("gender", tr('性别'), 70),
            ("cls", tr('班级'), 110),
            ("deleted_at", tr('删除时间'), 170),
        ):
            self.tree.heading(col_id, text=col_name)
            self.tree.column(col_id, width=dpi.scale(width), anchor="center")
        vsb = ttk.Scrollbar(table_frame, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(table_frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        self.tree.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")
        hsb.pack(side="bottom", fill="x")

        self.status_var = tk.StringVar(value="")
        ttk.Label(
            self, textvariable=self.status_var, padding=dpi.P((14, 6)),
            style="Muted.TLabel",
        ).pack(side="bottom", fill="x")

    def reload(self):
        try:
            rows = self.storage.list_deleted_students()
        except Exception as exc:
            messagebox.showerror(tr('存储错误'), str(exc), parent=self)
            return
        self.tree.delete(*self.tree.get_children())
        for index, row in enumerate(rows):
            self.tree.insert(
                "",
                "end",
                iid=row["student_id"],
                tags=("even" if index % 2 else "odd",),
                values=(
                    row["student_id"],
                    row["name"],
                    tr(row["gender"]),
                    row.get("class_name") or "",
                    str(row.get("deleted_at"))[:19],
                ),
            )
        zebra(self.tree)
        self.status_var.set(tr('回收站共 {} 名学生').format(len(rows)))

    def on_restore(self):
        selected = self.tree.selection()
        if not selected:
            messagebox.showwarning(tr('提示'), tr('请先选择要恢复的学生'), parent=self)
            return
        op, role = _operator(self)
        for sid in selected:
            try:
                if self.storage.restore_student(sid):
                    _log(self.storage, op, role, tr('恢复学生'), sid)
            except Exception as exc:
                messagebox.showerror(tr('存储错误'), str(exc), parent=self)
        self.reload()

    def on_purge(self):
        selected = self.tree.selection()
        if not selected:
            messagebox.showwarning(tr('提示'), tr('请先选择要彻底删除的学生'), parent=self)
            return
        if not messagebox.askyesno(
            tr('确认删除'),
        tr('确定彻底删除 {} 名学生吗？\n该学生的成绩将一并删除，且无法恢复。').format(len(selected)),
            parent=self,
        ):
            return
        op, role = _operator(self)
        for sid in selected:
            try:
                if self.storage.purge_student(sid):
                    _log(self.storage, op, role, tr('彻底删除学生'), sid)
            except Exception as exc:
                messagebox.showerror(tr('存储错误'), str(exc), parent=self)
        self.reload()
