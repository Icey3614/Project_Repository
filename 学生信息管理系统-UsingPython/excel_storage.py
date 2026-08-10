# -*- coding: utf-8 -*-
"""Excel 存储实现：在用户指定目录下创建/读取 students.xlsx、scores.xlsx、teachers.xlsx。"""
import os
import time

from openpyxl import Workbook, load_workbook

import config
from storage import (
    Storage,
    StorageError,
    default_student_password,
    default_teacher_password,
    hash_password,
    mask_id_card,
    validate_new_password,
    validate_score,
    validate_student,
    validate_teacher,
    verify_password,
)
from i18n import tr

STUDENT_HEADERS = [
    "学号", "姓名", "性别", "年龄", "入学年份", "身份证号", "密码哈希",
    "班级", "已删除", "删除时间",
]
SCORE_HEADERS = ["学号", "年份", "语文", "数学", "英语"]
TEACHER_HEADERS = ["账号", "密码哈希", "姓名", "角色", "工号", "身份证号"]
LOG_HEADERS = ["时间", "操作人", "角色", "操作", "对象", "详情"]

LOCK_STALE_SECONDS = 300  # 锁文件超过 5 分钟视为残留，允许接管
PURGE_DAYS = 30  # 回收站自动清空天数


def _to_int(value):
    if value is None or str(value).strip() == "":
        return None
    return int(value)


def _to_float(value):
    if value is None or str(value).strip() == "":
        return None
    return float(value)


class ExcelStorage(Storage):
    """使用 Excel 文件存储：students.xlsx / scores.xlsx / teachers.xlsx。"""

    mode = tr('Excel 表格')

    def __init__(self, directory):
        self.directory = os.path.abspath(directory)
        try:
            os.makedirs(self.directory, exist_ok=True)
        except OSError as exc:
            raise StorageError(tr('无法创建目录 {}：{}').format(self.directory, exc)) from exc
        self.students_file = os.path.join(self.directory, "students.xlsx")
        self.scores_file = os.path.join(self.directory, "scores.xlsx")
        self.teachers_file = os.path.join(self.directory, "teachers.xlsx")
        self.logs_file = os.path.join(self.directory, "logs.xlsx")
        self.lock_path = os.path.join(self.directory, "sms.lock")

        self._acquire_lock()

        self.students = {}  # student_id -> dict
        self.scores = {}    # (student_id, year) -> dict
        self.teachers = {}  # username -> dict
        self.logs = []      # list of dict
        self.first_use = False
        self.info = ""
        self._load_or_create()

    def _acquire_lock(self):
        """创建独占锁文件，防止多个程序同时操作同一份 Excel 数据。"""
        try:
            fd = os.open(self.lock_path, os.O_CREAT | os.O_EXCL | os.O_WRONLY)
            os.write(fd, str(os.getpid()).encode("utf-8"))
            os.close(fd)
        except FileExistsError:
            try:
                if time.time() - os.path.getmtime(self.lock_path) > LOCK_STALE_SECONDS:
                    os.remove(self.lock_path)
                    return self._acquire_lock()
            except OSError:
                pass
            raise StorageError(
                tr('该目录的数据文件正被另一个程序使用，请先关闭其他程序实例')
            )

    def release(self):
        """释放锁文件（程序退出时调用）。"""
        try:
            if os.path.exists(self.lock_path):
                os.remove(self.lock_path)
        except OSError:
            pass

    # ------------------------------------------------------------ 初始化

    def _load_or_create(self):
        had_any = False
        if os.path.exists(self.students_file):
            self._read_students()
            had_any = had_any or bool(self.students)
        if os.path.exists(self.scores_file):
            self._read_scores()
            had_any = had_any or bool(self.scores)
        if os.path.exists(self.teachers_file):
            self._read_teachers()
            had_any = had_any or bool(self.teachers)
        if os.path.exists(self.logs_file):
            self._read_logs()

        # 创建缺失的数据文件（含表头）
        self._write_students()
        self._write_scores()
        self._write_teachers()
        self._write_logs()

        # 首次使用：初始化默认管理员账号
        if not self.teachers:
            self.teachers[config.DEFAULT_TEACHER_USERNAME] = {
                "username": config.DEFAULT_TEACHER_USERNAME,
                "password_hash": hash_password(config.DEFAULT_TEACHER_PASSWORD),
                "real_name": config.DEFAULT_TEACHER_NAME,
                "role": "admin",
                "employee_id": "",
                "id_card": "",
            }
            self._write_teachers()

        self.first_use = not had_any
        # 回收站自动清空：删除超过 30 天的学生彻底清除
        self._purge_expired()
        if self.first_use:
            self.info = (
                tr('首次使用：已在 {} 创建 Excel 数据文件，并初始化默认教师账号 {}').format(self.directory, config.DEFAULT_TEACHER_USERNAME)
            )
        else:
            self.info = (
                tr('检测到已有数据：读取 {} 名学生、{} 条成绩、{} 个教师账号').format(len(self.students), len(self.scores), len(self.teachers))
            )

    def _read_students(self):
        try:
            wb = load_workbook(self.students_file, read_only=False)
            ws = wb.active
            header = [
                str(c).strip() if c is not None else ""
                for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            ]
            has_id = "身份证号" in header
            has_pw = "密码哈希" in header
            has_class = "班级" in header
            has_deleted = "已删除" in header
            has_deleted_at = "删除时间" in header
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or row[0] in (None, ""):
                    continue
                sid = str(row[0]).strip()
                if not sid:
                    continue
                deleted = 1 if has_deleted and len(row) > 8 and str(row[8] or "0").strip() in ("1", tr('是'), "True", "true") else 0
                self.students[sid] = {
                    "student_id": sid,
                    "name": str(row[1] or "").strip(),
                    "gender": str(row[2] or tr('男')).strip(),
                    "age": _to_int(row[3]),
                    "enroll_year": _to_int(row[4]),
                    "id_card": str(row[5] or "").strip() if has_id and len(row) > 5 else "",
                    "password_hash": str(row[6] or "") if has_pw and len(row) > 6 else "",
                    "class_name": str(row[7] or "").strip() if has_class and len(row) > 7 else "",
                    "deleted": deleted,
                    "deleted_at": row[9] if has_deleted_at and len(row) > 9 else None,
                }
            wb.close()
        except Exception as exc:
            raise StorageError(tr('读取学生表失败：{}').format(exc)) from exc

    def _read_scores(self):
        try:
            wb = load_workbook(self.scores_file)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or row[0] in (None, "") or row[1] in (None, ""):
                    continue
                sid = str(row[0]).strip()
                year = _to_int(row[1])
                if not sid or year is None:
                    continue
                self.scores[(sid, year)] = {
                    "student_id": sid,
                    "year": year,
                    "chinese_score": _to_float(row[2]),
                    "math_score": _to_float(row[3]),
                    "english_score": _to_float(row[4]),
                }
            wb.close()
        except Exception as exc:
            raise StorageError(tr('读取成绩表失败：{}').format(exc)) from exc

    def _read_teachers(self):
        try:
            wb = load_workbook(self.teachers_file)
            ws = wb.active
            header = [str(c).strip() if c is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
            has_role = "角色" in header
            has_employee = "工号" in header
            has_id_card = "身份证号" in header
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or row[0] in (None, ""):
                    continue
                username = str(row[0]).strip()
                if not username:
                    continue
                if has_role:
                    role = str(row[3] or "").strip()
                    if role not in ("teacher", "admin"):
                        role = "admin" if username == config.DEFAULT_TEACHER_USERNAME else "teacher"
                else:
                    role = "admin" if username == config.DEFAULT_TEACHER_USERNAME else "teacher"
                self.teachers[username] = {
                    "username": username,
                    "password_hash": str(row[1] or ""),
                    "real_name": str(row[2] or "").strip() or None,
                    "role": role,
                    "employee_id": str(row[4] or "").strip() if has_employee and len(row) > 4 else "",
                    "id_card": str(row[5] or "").strip() if has_id_card and len(row) > 5 else "",
                }
            wb.close()
        except Exception as exc:
            raise StorageError(tr('读取教师表失败：{}').format(exc)) from exc

    def _write_students(self):
        wb = Workbook()
        ws = wb.active
        ws.title = tr('学生')
        ws.append(STUDENT_HEADERS)
        for s in sorted(self.students.values(), key=lambda x: x["student_id"]):
            ws.append(
                [
                    s["student_id"],
                    s["name"],
                    s["gender"],
                    s["age"],
                    s["enroll_year"],
                    s.get("id_card", ""),
                    s.get("password_hash", ""),
                    s.get("class_name", "") or "",
                    1 if s.get("deleted") else 0,
                    s.get("deleted_at"),
                ]
            )
        self._autofit(ws, STUDENT_HEADERS)
        try:
            wb.save(self.students_file)
        except OSError as exc:
            raise StorageError(tr('保存学生表失败：{}').format(exc)) from exc
        finally:
            wb.close()

    def _write_scores(self):
        wb = Workbook()
        ws = wb.active
        ws.title = tr('成绩')
        ws.append(SCORE_HEADERS)
        for key in sorted(self.scores.keys(), key=lambda k: (k[0], k[1])):
            s = self.scores[key]
            ws.append(
                [
                    s["student_id"],
                    s["year"],
                    s["chinese_score"],
                    s["math_score"],
                    s["english_score"],
                ]
            )
        self._autofit(ws, SCORE_HEADERS)
        try:
            wb.save(self.scores_file)
        except OSError as exc:
            raise StorageError(tr('保存成绩表失败：{}').format(exc)) from exc
        finally:
            wb.close()

    def _write_teachers(self):
        wb = Workbook()
        ws = wb.active
        ws.title = tr('教师')
        ws.append(TEACHER_HEADERS)
        for t in sorted(self.teachers.values(), key=lambda x: x["username"]):
            ws.append(
                [
                    t["username"],
                    t["password_hash"],
                    t["real_name"],
                    t["role"],
                    t.get("employee_id", ""),
                    t.get("id_card", ""),
                ]
            )
        self._autofit(ws, TEACHER_HEADERS)
        try:
            wb.save(self.teachers_file)
        except OSError as exc:
            raise StorageError(tr('保存教师表失败：{}').format(exc)) from exc
        finally:
            wb.close()

    def _read_logs(self):
        try:
            wb = load_workbook(self.logs_file)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or row[0] in (None, ""):
                    continue
                self.logs.append(
                    {
                        "created_at": row[0],
                        "operator": str(row[1] or ""),
                        "role": str(row[2] or ""),
                        "action": str(row[3] or ""),
                        "target": str(row[4] or "") if len(row) > 4 else "",
                        "detail": str(row[5] or "") if len(row) > 5 else "",
                    }
                )
            wb.close()
        except Exception as exc:
            raise StorageError(tr('读取日志表失败：{}').format(exc)) from exc

    def _write_logs(self):
        wb = Workbook()
        ws = wb.active
        ws.title = tr('日志')
        ws.append(LOG_HEADERS)
        for r in self.logs[:500]:
            ws.append(
                [
                    r.get("created_at", ""),
                    r.get("operator", ""),
                    r.get("role", ""),
                    r.get("action", ""),
                    r.get("target", ""),
                    r.get("detail", ""),
                ]
            )
        try:
            wb.save(self.logs_file)
        except OSError as exc:
            raise StorageError(tr('保存日志表失败：{}').format(exc)) from exc
        finally:
            wb.close()

    def _purge_expired(self):
        """彻底删除回收站中超过 30 天的学生及其成绩。"""
        import datetime

        cutoff = datetime.datetime.now() - datetime.timedelta(days=PURGE_DAYS)
        expired = [
            sid
            for sid, s in self.students.items()
            if s.get("deleted") and s.get("deleted_at") is not None
            and s["deleted_at"] < cutoff
        ]
        for sid in expired:
            self.students.pop(sid, None)
            for key in [k for k in self.scores if k[0] == sid]:
                del self.scores[key]
        if expired:
            self._write_students()
            self._write_scores()

    @staticmethod
    def _autofit(ws, headers):
        for index, header in enumerate(headers, start=1):
            width = max(len(header) * 2 + 2, 12)
            if header in ("密码哈希",):
                width = 52
            ws.column_dimensions[ws.cell(row=1, column=index).column_letter].width = width

    def describe(self):
        return tr('Excel 表格（{}）').format(self.directory)

    # ------------------------------------------------------------ 学生

    def get_student(self, student_id):
        row = self.students.get(student_id)
        if row and row.get("deleted"):
            return None
        return row

    def list_students(self, keyword=None):
        rows = [r for r in self.students.values() if not r.get("deleted")]
        if keyword:
            keyword = keyword.strip()
            rows = [
                r
                for r in rows
                if keyword in r["student_id"] or keyword in r["name"]
            ]
        return sorted(rows, key=lambda r: r["student_id"])

    def list_deleted_students(self):
        rows = [r for r in self.students.values() if r.get("deleted")]
        return sorted(rows, key=lambda r: r.get("deleted_at") or "", reverse=True)

    def student_login(self, student_id, password):
        row = self.students.get(student_id.strip())
        if (
            row
            and row.get("password_hash")
            and verify_password(password, row["password_hash"])
        ):
            return {
                "student_id": row["student_id"],
                "name": row["name"],
                "gender": row["gender"],
                "age": row["age"],
                "enroll_year": row["enroll_year"],
                "id_card": row.get("id_card", ""),
                "class_name": row.get("class_name", ""),
            }
        return None

    def add_student(
        self, student_id, name, gender, age, enroll_year, id_card,
        password=None, class_name="",
    ):
        errors = validate_student(
            student_id, name, gender, age, enroll_year, id_card, class_name
        )
        if errors:
            raise ValueError("；".join(errors))
        sid = student_id.strip()
        if sid in self.students:
            raise ValueError(tr('学号 {} 已存在，请勿重复添加').format(sid))
        if password is None:
            password = default_student_password(sid, id_card)
        self.students[sid] = {
            "student_id": sid,
            "name": name.strip(),
            "gender": gender,
            "age": _to_int(age),
            "enroll_year": _to_int(enroll_year),
            "id_card": str(id_card).strip(),
            "password_hash": hash_password(password),
            "class_name": str(class_name).strip() or "",
            "deleted": 0,
            "deleted_at": None,
        }
        self._write_students()
        return password

    def update_student(
        self, student_id, name, gender, age, enroll_year, id_card, class_name=""
    ):
        errors = validate_student(
            student_id, name, gender, age, enroll_year, id_card, class_name
        )
        if errors:
            raise ValueError("；".join(errors))
        sid = student_id.strip()
        if sid not in self.students:
            raise ValueError(tr('学号 {} 不存在，无法修改').format(sid))
        generated = None
        if not self.students[sid].get("password_hash"):
            generated = default_student_password(sid, id_card)
            self.students[sid]["password_hash"] = hash_password(generated)
        self.students[sid].update(
            {
                "name": name.strip(),
                "gender": gender,
                "age": _to_int(age),
                "enroll_year": _to_int(enroll_year),
                "id_card": str(id_card).strip(),
                "class_name": str(class_name).strip() or "",
            }
        )
        self._write_students()
        return generated

    def change_student_password(self, student_id, old_password, new_password):
        strength_error = validate_new_password(new_password)
        if strength_error:
            raise ValueError(strength_error)
        row = self.students.get(student_id.strip())
        if not row or not row.get("password_hash"):
            raise ValueError(tr('该学生还没有登录密码，请联系老师设置'))
        if not verify_password(old_password, row["password_hash"]):
            raise ValueError(tr('原密码不正确'))
        row["password_hash"] = hash_password(new_password)
        self._write_students()

    def delete_student(self, student_id):
        row = self.students.get(student_id)
        if row is None or row.get("deleted"):
            return False
        row["deleted"] = 1
        import datetime
        row["deleted_at"] = datetime.datetime.now()
        self._write_students()
        return True

    def restore_student(self, student_id):
        row = self.students.get(student_id)
        if row is None or not row.get("deleted"):
            return False
        row["deleted"] = 0
        row["deleted_at"] = None
        self._write_students()
        return True

    def purge_student(self, student_id):
        if self.students.pop(student_id, None) is None:
            return False
        for key in [k for k in self.scores if k[0] == student_id]:
            del self.scores[key]
        self._write_students()
        self._write_scores()
        return True

    # ------------------------------------------------------------ 成绩

    def get_scores(self, student_id):
        rows = [
            self.scores[k]
            for k in sorted(self.scores.keys(), key=lambda k: (k[0], k[1]))
            if k[0] == student_id
        ]
        return rows

    def get_all_scores(self):
        rows = [
            self.scores[k]
            for k in sorted(self.scores.keys(), key=lambda k: (k[0], k[1]))
        ]
        return rows

    def get_score(self, student_id, year):
        return self.scores.get((student_id, year))

    def add_score(self, student_id, year, chinese, math, english):
        errors = validate_score(year, chinese, math, english)
        if errors:
            raise ValueError("；".join(errors))
        if student_id not in self.students:
            raise ValueError(tr('该学生不存在，无法登记成绩'))
        key = (student_id, int(year))
        if key in self.scores:
            raise ValueError(tr('{} 年的成绩已登记，如需修改请使用“修改”功能').format(year))
        self.scores[key] = {
            "student_id": student_id,
            "year": int(year),
            "chinese_score": _to_float(chinese),
            "math_score": _to_float(math),
            "english_score": _to_float(english),
        }
        self._write_scores()

    def update_score(self, student_id, year, chinese, math, english):
        errors = validate_score(year, chinese, math, english)
        if errors:
            raise ValueError("；".join(errors))
        key = (student_id, int(year))
        if key not in self.scores:
            raise ValueError(tr('{} 年的成绩记录不存在，无法修改').format(year))
        self.scores[key].update(
            {
                "chinese_score": _to_float(chinese),
                "math_score": _to_float(math),
                "english_score": _to_float(english),
            }
        )
        self._write_scores()

    def delete_score(self, student_id, year):
        if self.scores.pop((student_id, year), None) is None:
            return False
        self._write_scores()
        return True

    # ------------------------------------------------------------ 教师

    def _login(self, username, password, role):
        row = self.teachers.get(username.strip())
        if row and row.get("role") == role and verify_password(password, row["password_hash"]):
            return {
                "teacher_id": username,
                "username": row["username"],
                "real_name": row["real_name"],
                "role": row["role"],
            }
        return None

    def teacher_login(self, username, password):
        return self._login(username, password, "teacher")

    def admin_login(self, username, password):
        return self._login(username, password, "admin")

    def change_teacher_password(self, username, old_password, new_password):
        strength_error = validate_new_password(new_password)
        if strength_error:
            raise ValueError(strength_error)
        row = self.teachers.get(username.strip())
        if not row or not verify_password(old_password, row["password_hash"]):
            raise ValueError(tr('原密码不正确'))
        row["password_hash"] = hash_password(new_password)
        self._write_teachers()

    def add_teacher(
        self, username, real_name, password, role="teacher",
        employee_id="", id_card="",
    ):
        errors = validate_teacher(username, employee_id, id_card, role)
        if errors:
            raise ValueError("；".join(errors))
        if password is None:
            password = default_teacher_password(employee_id, id_card)
        uname = username.strip()
        if uname in self.teachers:
            raise ValueError(tr('教师账号 {} 已存在').format(uname))
        self.teachers[uname] = {
            "username": uname,
            "password_hash": hash_password(password),
            "real_name": str(real_name).strip() or None,
            "role": role,
            "employee_id": str(employee_id).strip(),
            "id_card": str(id_card).strip(),
        }
        self._write_teachers()
        return password

    def list_teachers(self):
        rows = list(self.teachers.values())
        return sorted(
            rows, key=lambda r: (r.get("role") != "admin", r["username"])
        )

    def update_teacher(
        self, username, real_name, role, new_password=None,
        employee_id=None, id_card=None,
    ):
        if role not in ("teacher", "admin"):
            raise ValueError(tr('角色只能是“教师”或“管理员”'))
        uname = username.strip()
        row = self.teachers.get(uname)
        if row is None:
            raise ValueError(tr('教师账号 {} 不存在').format(uname))
        if new_password:
            strength_error = validate_new_password(new_password)
            if strength_error:
                raise ValueError(strength_error)
            row["password_hash"] = hash_password(new_password)
        row["real_name"] = str(real_name).strip() or None
        row["role"] = role
        if employee_id is not None:
            row["employee_id"] = str(employee_id).strip()
        if id_card is not None:
            row["id_card"] = str(id_card).strip()
        self._write_teachers()

    def delete_teacher(self, username):
        if self.teachers.pop(username.strip(), None) is None:
            return False
        self._write_teachers()
        return True

    # ------------------------------------------------------------ 日志与导出

    def add_log(self, operator, role, action, target="", detail=""):
        import datetime

        self.logs.insert(
            0,
            {
                "created_at": datetime.datetime.now(),
                "operator": str(operator)[:50],
                "role": str(role)[:20],
                "action": str(action)[:50],
                "target": str(target)[:100],
                "detail": str(detail)[:255],
            },
        )
        self._write_logs()

    def list_logs(self, limit=500):
        return self.logs[: int(limit)]

    def clear_logs(self):
        """清空全部操作日志（仅管理员调用）。"""
        self.logs.clear()
        self._write_logs()

    def export_to_excel(self, path):
        """把学生、成绩、教师、日志全部导出到一个 Excel 文件。"""
        wb = Workbook()
        ws1 = wb.active
        ws1.title = tr('学生')
        ws1.append(
            ["学号", "姓名", "性别", "年龄", "入学年份", "身份证号", "班级", "状态"]
        )
        for s in self.list_students():
            ws1.append(
                [
                    s["student_id"], s["name"], s["gender"], s["age"],
                    s["enroll_year"], s.get("id_card", ""),
                    s.get("class_name", ""), "在读",
                ]
            )
        for s in self.list_deleted_students():
            ws1.append(
                [
                    s["student_id"], s["name"], s["gender"], s["age"],
                    s["enroll_year"], s.get("id_card", ""),
                    s.get("class_name", ""), "已删除",
                ]
            )
        ws2 = wb.create_sheet(tr('成绩'))
        ws2.append(["学号", "年份", "语文", "数学", "英语"])
        for r in self.get_all_scores():
            ws2.append(
                [
                    r["student_id"], r["year"],
                    r["chinese_score"], r["math_score"], r["english_score"],
                ]
            )
        ws3 = wb.create_sheet(tr('教师'))
        ws3.append(["账号", "姓名", "角色", "工号", "身份证号"])
        for t in self.list_teachers():
            ws3.append(
                [
                    t["username"], t.get("real_name") or "",
                    tr('管理员') if t.get("role") == "admin" else tr('教师'),
                    t.get("employee_id") or "", t.get("id_card") or "",
                ]
            )
        ws4 = wb.create_sheet(tr('日志'))
        ws4.append(["时间", "操作人", "角色", "操作", "对象", "详情"])
        for r in self.list_logs(500):
            ws4.append(
                [
                    r.get("created_at", ""), r.get("operator", ""),
                    r.get("role", ""), r.get("action", ""),
                    r.get("target", ""), r.get("detail", ""),
                ]
            )
        wb.save(path)
        wb.close()
