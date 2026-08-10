# -*- coding: utf-8 -*-
"""MySQL 存储实现：自动创建项目专属数据库与数据表，并区分首次/多次使用。"""
import pymysql
import pymysql.cursors

import config
from storage import (
    Storage,
    StorageError,
    default_student_password,
    default_teacher_password,
    hash_password,
    mask_id_card,
    validate_new_password,
    validate_score,
    validate_student,
    validate_teacher,
    verify_password,
)
from i18n import tr


SCHEMA_SQL = [
    """
CREATE TABLE IF NOT EXISTS students (
  student_id VARCHAR(20) NOT NULL COMMENT '学号',
  name VARCHAR(50) NOT NULL COMMENT '姓名',
  gender ENUM('男','女') NOT NULL COMMENT '性别',
  age INT NULL COMMENT '年龄',
  enroll_year INT NULL COMMENT '入学年份',
  id_card VARCHAR(18) NULL COMMENT '身份证号',
  password_hash VARCHAR(255) NULL COMMENT '登录密码哈希',
  class_name VARCHAR(50) NULL COMMENT '班级',
  deleted TINYINT(1) NOT NULL DEFAULT 0 COMMENT '是否已删除（软删除）',
  deleted_at DATETIME NULL COMMENT '删除时间',
  created_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP,
  updated_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
  PRIMARY KEY (student_id)
) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_0900_ai_ci COMMENT='学生信息表';
""",
    """
CREATE TABLE IF NOT EXISTS student_scores (
  student_id VARCHAR(20) NOT NULL COMMENT '学号',
  year INT NOT NULL COMMENT '学年（如 2025）',
  chinese_score DECIMAL(5,1) NULL COMMENT '语文成绩',
  math_score DECIMAL(5,1) NULL COMMENT '数学成绩',
  english_score DECIMAL(5,1) NULL COMMENT '英语成绩',
  PRIMARY KEY (student_id, year),
  CONSTRAINT fk_scores_student FOREIGN KEY (student_id)
    REFERENCES students (student_id) ON DELETE CASCADE
) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_0900_ai_ci COMMENT='学生分年度成绩表';
""",
    """
CREATE TABLE IF NOT EXISTS teachers (
  teacher_id INT NOT NULL AUTO_INCREMENT,
  username VARCHAR(50) NOT NULL COMMENT '登录账号',
  password_hash VARCHAR(255) NOT NULL COMMENT '密码哈希',
  real_name VARCHAR(50) NULL COMMENT '姓名',
  role ENUM('teacher','admin') NOT NULL DEFAULT 'teacher' COMMENT '角色',
  employee_id VARCHAR(20) NULL COMMENT '工号',
  id_card VARCHAR(18) NULL COMMENT '身份证号',
  created_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP,
  PRIMARY KEY (teacher_id),
  UNIQUE KEY uk_username (username)
) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_0900_ai_ci COMMENT='教师账号表';
""",
    """
CREATE TABLE IF NOT EXISTS logs (
  log_id BIGINT NOT NULL AUTO_INCREMENT,
  created_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP,
  operator VARCHAR(50) NOT NULL COMMENT '操作人',
  role VARCHAR(20) NOT NULL COMMENT '操作人角色',
  action VARCHAR(50) NOT NULL COMMENT '操作',
  target VARCHAR(100) NULL COMMENT '对象',
  detail VARCHAR(255) NULL COMMENT '详情',
  PRIMARY KEY (log_id),
  KEY idx_created_at (created_at)
) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_0900_ai_ci COMMENT='操作日志表';
""",
]


def _to_int_or_none(value):
    if value is None or str(value).strip() == "":
        return None
    return int(value)


def _to_float_or_none(value):
    if value is None or str(value).strip() == "":
        return None
    return float(value)


def _column_exists(cur, table, column):
    cur.execute(
        "SELECT COUNT(*) AS c FROM information_schema.COLUMNS "
        "WHERE TABLE_SCHEMA = DATABASE() AND TABLE_NAME = %s AND COLUMN_NAME = %s",
        (table, column),
    )
    return cur.fetchone()["c"] > 0


class MySqlStorage(Storage):
    """使用本机 MySQL 数据库存储，自动创建项目专属数据库。"""

    mode = tr('MySQL 数据库')

    def __init__(self, host, port, user, password):
        self.host = host
        self.port = int(port)
        self.user = user
        self.password = password
        self.database = config.DATABASE_NAME
        self.first_use = False
        self.info = ""
        self._ensure()

    # ------------------------------------------------------------ 初始化

    def _connect(self, database=None):
        # 断线重连：连接失败时短暂等待后重试一次
        last_exc = None
        for attempt in range(2):
            try:
                return pymysql.connect(
                    host=self.host,
                    port=self.port,
                    user=self.user,
                    password=self.password,
                    database=database,
                    charset="utf8mb4",
                    cursorclass=pymysql.cursors.DictCursor,
                    autocommit=False,
                )
            except pymysql.err.OperationalError as exc:
                last_exc = exc
                if attempt == 0:
                    import time
                    time.sleep(1.0)
        raise StorageError(tr('无法连接 MySQL：{}').format(last_exc)) from last_exc

    def _ensure(self):
        """连接服务器 → 创建项目专属数据库 → 建表 → 初始化默认账号。"""
        # 第一步：连接服务器（不指定数据库），检查/创建项目专属数据库
        conn = self._connect()
        db_existed = True
        try:
            with conn.cursor() as cur:
                cur.execute(
                    "SELECT SCHEMA_NAME FROM information_schema.SCHEMATA "
                    "WHERE SCHEMA_NAME = %s",
                    (self.database,),
                )
                db_existed = cur.fetchone() is not None
                if not db_existed:
                    cur.execute(
                        f"CREATE DATABASE `{self.database}` CHARACTER SET utf8mb4"
                    )
            conn.commit()
        finally:
            conn.close()

        # 第二步：连接项目专属数据库并建表
        conn = self._connect(self.database)
        try:
            with conn.cursor() as cur:
                for statement in SCHEMA_SQL:
                    cur.execute(statement)

                # 旧版本补充学生身份证号/登录密码列
                if not _column_exists(cur, "students", "id_card"):
                    cur.execute(
                        "ALTER TABLE students ADD COLUMN id_card VARCHAR(18) NULL COMMENT '身份证号' AFTER enroll_year"
                    )
                if not _column_exists(cur, "students", "password_hash"):
                    cur.execute(
                        "ALTER TABLE students ADD COLUMN password_hash VARCHAR(255) NULL COMMENT '登录密码哈希' AFTER id_card"
                    )
                if not _column_exists(cur, "students", "class_name"):
                    cur.execute(
                        "ALTER TABLE students ADD COLUMN class_name VARCHAR(50) NULL COMMENT '班级' AFTER password_hash"
                    )
                if not _column_exists(cur, "students", "deleted"):
                    cur.execute(
                        "ALTER TABLE students ADD COLUMN deleted TINYINT(1) NOT NULL DEFAULT 0 COMMENT '是否已删除（软删除）' AFTER class_name"
                    )
                if not _column_exists(cur, "students", "deleted_at"):
                    cur.execute(
                        "ALTER TABLE students ADD COLUMN deleted_at DATETIME NULL COMMENT '删除时间' AFTER deleted"
                    )
                # 旧版本补充教师工号/身份证号列
                if not _column_exists(cur, "teachers", "employee_id"):
                    cur.execute(
                        "ALTER TABLE teachers ADD COLUMN employee_id VARCHAR(20) NULL COMMENT '工号' AFTER role"
                    )
                if not _column_exists(cur, "teachers", "id_card"):
                    cur.execute(
                        "ALTER TABLE teachers ADD COLUMN id_card VARCHAR(18) NULL COMMENT '身份证号' AFTER employee_id"
                    )

                # 旧版本 teachers 表没有 role 列 → 补充并把默认账号设为管理员
                if not _column_exists(cur, "teachers", "role"):
                    cur.execute(
                        "ALTER TABLE teachers ADD COLUMN role ENUM('teacher','admin') NOT NULL DEFAULT 'teacher' COMMENT '角色' AFTER real_name"
                    )
                    cur.execute(
                        "UPDATE teachers SET role = 'admin' WHERE username = %s",
                        (config.DEFAULT_TEACHER_USERNAME,),
                    )

                # 默认管理员账号（仅当教师表为空时创建）
                cur.execute("SELECT COUNT(*) AS c FROM teachers")
                if cur.fetchone()["c"] == 0:
                    cur.execute(
                        "INSERT INTO teachers "
                        "(username, password_hash, real_name, role) "
                        "VALUES (%s, %s, %s, 'admin')",
                        (
                            config.DEFAULT_TEACHER_USERNAME,
                            hash_password(config.DEFAULT_TEACHER_PASSWORD),
                            config.DEFAULT_TEACHER_NAME,
                        ),
                    )
                cur.execute("SELECT COUNT(*) AS c FROM students")
                student_count = cur.fetchone()["c"]
                # 回收站自动清空：删除超过 30 天的学生彻底清除
                cur.execute(
                    "DELETE FROM students WHERE deleted = 1 "
                    "AND deleted_at < NOW() - INTERVAL 30 DAY"
                )
            conn.commit()
        finally:
            conn.close()

        # 首次使用与多次使用的判断：是否有学生数据
        self.first_use = student_count == 0
        if self.first_use:
            self.info = (
                tr('首次使用：已创建项目专属数据库 {} 及数据表，并初始化默认教师账号 {}').format(self.database, config.DEFAULT_TEACHER_USERNAME)
            )
        else:
            self.info = (
                tr('检测到已有数据：数据库 {} 中现有 {} 名学生，将直接读取使用').format(self.database, student_count)
            )

    def describe(self):
        return tr('MySQL 数据库（{}）').format(self.database)

    # ------------------------------------------------------------ 学生

    def get_student(self, student_id):
        conn = self._connect(self.database)
        try:
            with conn.cursor() as cur:
                cur.execute(
                    "SELECT * FROM students WHERE student_id = %s AND deleted = 0",
                    (student_id,),
                )
                return cur.fetchone()
        finally:
            conn.close()

    def list_students(self, keyword=None):
        conn = self._connect(self.database)
        try:
            with conn.cursor() as cur:
                if keyword:
                    like = f"%{keyword.strip()}%"
                    cur.execute(
                        "SELECT * FROM students WHERE deleted = 0 "
                        "AND (student_id LIKE %s OR name LIKE %s) ORDER BY student_id",
                        (like, like),
                    )
                else:
                    cur.execute(
                        "SELECT * FROM students WHERE deleted = 0 ORDER BY student_id"
                    )
                return cur.fetchall()
        finally:
            conn.close()

    def list_deleted_students(self):
        conn = self._connect(self.database)
        try:
            with conn.cursor() as cur:
                cur.execute(
                    "SELECT * FROM students WHERE deleted = 1 "
                    "ORDER BY deleted_at DESC"
                )
                return cur.fetchall()
        finally:
            conn.close()

    def student_login(self, student_id, password):
        """学生登录：学号 + 密码。"""
        conn = self._connect(self.database)
        try:
            with conn.cursor() as cur:
                cur.execute(
                    "SELECT * FROM students WHERE student_id = %s AND deleted = 0",
                    (student_id.strip(),),
                )
                row = cur.fetchone()
            if (
                row
                and row.get("password_hash")
                and verify_password(password, row["password_hash"])
            ):
                return {
                    "student_id": row["student_id"],
                    "name": row["name"],
                    "gender": row["gender"],
                    "age": row["age"],
                    "enroll_year": row["enroll_year"],
                    "id_card": row["id_card"],
                    "class_name": row["class_name"],
                }
            return None
        finally:
            conn.close()

    def add_student(
        self, student_id, name, gender, age, enroll_year, id_card,
        password=None, class_name="",
    ):
        errors = validate_student(
            student_id, name, gender, age, enroll_year, id_card, class_name
        )
        if errors:
            raise ValueError("；".join(errors))
        if password is None:
            password = default_student_password(student_id, id_card)
        conn = self._connect(self.database)
        try:
            with conn.cursor() as cur:
                try:
                    cur.execute(
                        "INSERT INTO students "
                        "(student_id, name, gender, age, enroll_year, id_card, "
                        "password_hash, class_name) "
                        "VALUES (%s, %s, %s, %s, %s, %s, %s, %s)",
                        (
                            student_id.strip(),
                            name.strip(),
                            gender,
                            _to_int_or_none(age),
                            _to_int_or_none(enroll_year),
                            str(id_card).strip(),
                            hash_password(password),
                            str(class_name).strip() or None,
                        ),
                    )
                except pymysql.err.IntegrityError:
                    raise ValueError(tr('学号 {} 已存在，请勿重复添加').format(student_id))
            conn.commit()
            return password
        finally:
            conn.close()

    def update_student(
        self, student_id, name, gender, age, enroll_year, id_card, class_name=""
    ):
        errors = validate_student(
            student_id, name, gender, age, enroll_year, id_card, class_name
        )
        if errors:
            raise ValueError("；".join(errors))
        conn = self._connect(self.database)
        try:
            with conn.cursor() as cur:
                cur.execute(
                    "SELECT password_hash FROM students "
                    "WHERE student_id = %s AND deleted = 0",
                    (student_id.strip(),),
                )
                current = cur.fetchone()
                if current is None:
                    raise ValueError(tr('学号 {} 不存在，无法修改').format(student_id))
                generated = None
                if not current.get("password_hash"):
                    # 学生还没有登录密码：用新身份证号生成初始密码
                    generated = default_student_password(student_id, id_card)
                affected = cur.execute(
                    "UPDATE students SET name=%s, gender=%s, age=%s, "
                    "enroll_year=%s, id_card=%s, class_name=%s, "
                    "password_hash=COALESCE(password_hash, %s) "
                    "WHERE student_id=%s AND deleted = 0",
                    (
                        name.strip(),
                        gender,
                        _to_int_or_none(age),
                        _to_int_or_none(enroll_year),
                        str(id_card).strip(),
                        str(class_name).strip() or None,
                        hash_password(generated) if generated else None,
                        student_id.strip(),
                    ),
                )
                if affected == 0:
                    raise ValueError(tr('学号 {} 不存在，无法修改').format(student_id))
            conn.commit()
            return generated
        finally:
            conn.close()

    def change_student_password(self, student_id, old_password, new_password):
        strength_error = validate_new_password(new_password)
        if strength_error:
            raise ValueError(strength_error)
        conn = self._connect(self.database)
        try:
            with conn.cursor() as cur:
                cur.execute(
                    "SELECT password_hash FROM students "
                    "WHERE student_id = %s AND deleted = 0",
                    (student_id.strip(),),
                )
                row = cur.fetchone()
                if not row or not row.get("password_hash"):
                    raise ValueError(tr('该学生还没有登录密码，请联系老师设置'))
                if not verify_password(old_password, row["password_hash"]):
                    raise ValueError(tr('原密码不正确'))
                cur.execute(
                    "UPDATE students SET password_hash = %s WHERE student_id = %s",
                    (hash_password(new_password), student_id.strip()),
                )
            conn.commit()
        finally:
            conn.close()

    def delete_student(self, student_id):
        conn = self._connect(self.database)
        try:
            with conn.cursor() as cur:
                affected = cur.execute(
                    "UPDATE students SET deleted = 1, deleted_at = NOW() "
                    "WHERE student_id = %s AND deleted = 0",
                    (student_id,),
                )
            conn.commit()
            return affected > 0
        finally:
            conn.close()

    def restore_student(self, student_id):
        conn = self._connect(self.database)
        try:
            with conn.cursor() as cur:
                affected = cur.execute(
                    "UPDATE students SET deleted = 0, deleted_at = NULL "
                    "WHERE student_id = %s AND deleted = 1",
                    (student_id,),
                )
            conn.commit()
            return affected > 0
        finally:
            conn.close()

    def purge_student(self, student_id):
        conn = self._connect(self.database)
        try:
            with conn.cursor() as cur:
                affected = cur.execute(
                    "DELETE FROM students WHERE student_id = %s", (student_id,)
                )
            conn.commit()
            return affected > 0
        finally:
            conn.close()

    # ------------------------------------------------------------ 成绩

    def get_scores(self, student_id):
        conn = self._connect(self.database)
        try:
            with conn.cursor() as cur:
                cur.execute(
                    "SELECT * FROM student_scores WHERE student_id = %s "
                    "ORDER BY year",
                    (student_id,),
                )
                return cur.fetchall()
        finally:
            conn.close()

    def get_all_scores(self):
        conn = self._connect(self.database)
        try:
            with conn.cursor() as cur:
                cur.execute(
                    "SELECT student_id, year, chinese_score, math_score, english_score "
                    "FROM student_scores ORDER BY year, student_id"
                )
                return cur.fetchall()
        finally:
            conn.close()

    def get_score(self, student_id, year):
        conn = self._connect(self.database)
        try:
            with conn.cursor() as cur:
                cur.execute(
                    "SELECT * FROM student_scores WHERE student_id = %s AND year = %s",
                    (student_id, year),
                )
                return cur.fetchone()
        finally:
            conn.close()

    def add_score(self, student_id, year, chinese, math, english):
        errors = validate_score(year, chinese, math, english)
        if errors:
            raise ValueError("；".join(errors))
        if self.get_student(student_id) is None:
            raise ValueError(tr('该学生不存在，无法登记成绩'))
        conn = self._connect(self.database)
        try:
            with conn.cursor() as cur:
                try:
                    cur.execute(
                        "INSERT INTO student_scores "
                        "(student_id, year, chinese_score, math_score, english_score) "
                        "VALUES (%s, %s, %s, %s, %s)",
                        (
                            student_id,
                            int(year),
                            _to_float_or_none(chinese),
                            _to_float_or_none(math),
                            _to_float_or_none(english),
                        ),
                    )
                except pymysql.err.IntegrityError:
                    raise ValueError(tr('{} 年的成绩已登记，如需修改请使用“修改”功能').format(year))
            conn.commit()
        finally:
            conn.close()

    def update_score(self, student_id, year, chinese, math, english):
        errors = validate_score(year, chinese, math, english)
        if errors:
            raise ValueError("；".join(errors))
        conn = self._connect(self.database)
        try:
            with conn.cursor() as cur:
                affected = cur.execute(
                    "UPDATE student_scores SET chinese_score=%s, math_score=%s, "
                    "english_score=%s WHERE student_id=%s AND year=%s",
                    (
                        _to_float_or_none(chinese),
                        _to_float_or_none(math),
                        _to_float_or_none(english),
                        student_id,
                        int(year),
                    ),
                )
                if affected == 0:
                    raise ValueError(tr('{} 年的成绩记录不存在，无法修改').format(year))
            conn.commit()
        finally:
            conn.close()

    def delete_score(self, student_id, year):
        conn = self._connect(self.database)
        try:
            with conn.cursor() as cur:
                affected = cur.execute(
                    "DELETE FROM student_scores WHERE student_id = %s AND year = %s",
                    (student_id, year),
                )
            conn.commit()
            return affected > 0
        finally:
            conn.close()

    # ------------------------------------------------------------ 教师

    def _login(self, username, password, role):
        conn = self._connect(self.database)
        try:
            with conn.cursor() as cur:
                cur.execute(
                    "SELECT * FROM teachers WHERE username = %s AND role = %s",
                    (username.strip(), role),
                )
                row = cur.fetchone()
            if row and verify_password(password, row["password_hash"]):
                return {
                    "teacher_id": row["teacher_id"],
                    "username": row["username"],
                    "real_name": row["real_name"],
                    "role": row["role"],
                }
            return None
        finally:
            conn.close()

    def teacher_login(self, username, password):
        return self._login(username, password, "teacher")

    def admin_login(self, username, password):
        return self._login(username, password, "admin")

    def change_teacher_password(self, username, old_password, new_password):
        strength_error = validate_new_password(new_password)
        if strength_error:
            raise ValueError(strength_error)
        conn = self._connect(self.database)
        try:
            with conn.cursor() as cur:
                cur.execute(
                    "SELECT * FROM teachers WHERE username = %s",
                    (username.strip(),),
                )
                row = cur.fetchone()
                if not row or not verify_password(old_password, row["password_hash"]):
                    raise ValueError(tr('原密码不正确'))
                cur.execute(
                    "UPDATE teachers SET password_hash = %s WHERE username = %s",
                    (hash_password(new_password), username.strip()),
                )
            conn.commit()
        finally:
            conn.close()

    def add_teacher(
        self, username, real_name, password, role="teacher",
        employee_id="", id_card="",
    ):
        errors = validate_teacher(username, employee_id, id_card, role)
        if errors:
            raise ValueError("；".join(errors))
        if password is None:
            password = default_teacher_password(employee_id, id_card)
        conn = self._connect(self.database)
        try:
            with conn.cursor() as cur:
                try:
                    cur.execute(
                        "INSERT INTO teachers "
                        "(username, password_hash, real_name, role, employee_id, id_card) "
                        "VALUES (%s, %s, %s, %s, %s, %s)",
                        (
                            username.strip(),
                            hash_password(password),
                            str(real_name).strip() or None,
                            role,
                            str(employee_id).strip(),
                            str(id_card).strip(),
                        ),
                    )
                except pymysql.err.IntegrityError:
                    raise ValueError(tr('教师账号 {} 已存在').format(username))
            conn.commit()
            return password
        finally:
            conn.close()

    def list_teachers(self):
        conn = self._connect(self.database)
        try:
            with conn.cursor() as cur:
                cur.execute(
                    "SELECT teacher_id, username, real_name, role, "
                    "employee_id, id_card "
                    "FROM teachers ORDER BY role DESC, username"
                )
                return cur.fetchall()
        finally:
            conn.close()

    def update_teacher(
        self, username, real_name, role, new_password=None,
        employee_id=None, id_card=None,
    ):
        if role not in ("teacher", "admin"):
            raise ValueError(tr('角色只能是“教师”或“管理员”'))
        conn = self._connect(self.database)
        try:
            with conn.cursor() as cur:
                if new_password:
                    strength_error = validate_new_password(new_password)
                    if strength_error:
                        raise ValueError(strength_error)
                    affected = cur.execute(
                        "UPDATE teachers SET real_name=%s, role=%s, "
                        "password_hash=%s, employee_id=%s, id_card=%s "
                        "WHERE username=%s",
                        (
                            str(real_name).strip() or None,
                            role,
                            hash_password(new_password),
                            str(employee_id).strip() if employee_id else None,
                            str(id_card).strip() if id_card else None,
                            username.strip(),
                        ),
                    )
                else:
                    affected = cur.execute(
                        "UPDATE teachers SET real_name=%s, role=%s, "
                        "employee_id=%s, id_card=%s "
                        "WHERE username=%s",
                        (
                            str(real_name).strip() or None,
                            role,
                            str(employee_id).strip() if employee_id else None,
                            str(id_card).strip() if id_card else None,
                            username.strip(),
                        ),
                    )
                if affected == 0:
                    raise ValueError(tr('教师账号 {} 不存在').format(username))
            conn.commit()
        finally:
            conn.close()

    def delete_teacher(self, username):
        conn = self._connect(self.database)
        try:
            with conn.cursor() as cur:
                affected = cur.execute(
                    "DELETE FROM teachers WHERE username = %s", (username.strip(),)
                )
            conn.commit()
            return affected > 0
        finally:
            conn.close()

    # ------------------------------------------------------------ 日志与导出

    def add_log(self, operator, role, action, target="", detail=""):
        conn = self._connect(self.database)
        try:
            with conn.cursor() as cur:
                cur.execute(
                    "INSERT INTO logs (operator, role, action, target, detail) "
                    "VALUES (%s, %s, %s, %s, %s)",
                    (
                        str(operator)[:50],
                        str(role)[:20],
                        str(action)[:50],
                        str(target)[:100],
                        str(detail)[:255],
                    ),
                )
            conn.commit()
        finally:
            conn.close()

    def list_logs(self, limit=500):
        conn = self._connect(self.database)
        try:
            with conn.cursor() as cur:
                cur.execute(
                    "SELECT log_id, created_at, operator, role, action, "
                    "target, detail FROM logs ORDER BY log_id DESC LIMIT %s",
                    (int(limit),),
                )
                return cur.fetchall()
        finally:
            conn.close()

    def clear_logs(self):
        """清空全部操作日志（仅管理员调用）。"""
        conn = self._connect(self.database)
        try:
            with conn.cursor() as cur:
                cur.execute("DELETE FROM logs")
            conn.commit()
        finally:
            conn.close()

    def export_to_excel(self, path):
        """把学生、成绩、教师、日志全部导出到一个 Excel 文件。"""
        from openpyxl import Workbook

        students = self.list_students()
        deleted = self.list_deleted_students()
        scores = self.get_all_scores()
        teachers = self.list_teachers()
        logs = self.list_logs(500)

        wb = Workbook()
        ws1 = wb.active
        ws1.title = tr('学生')
        ws1.append(
            ["学号", "姓名", "性别", "年龄", "入学年份", "身份证号", "班级", "状态"]
        )
        for s in students:
            ws1.append(
                [
                    s["student_id"], s["name"], s["gender"], s["age"],
                    s["enroll_year"], s["id_card"], s.get("class_name") or "",
                    "在读",
                ]
            )
        for s in deleted:
            ws1.append(
                [
                    s["student_id"], s["name"], s["gender"], s["age"],
                    s["enroll_year"], s["id_card"], s.get("class_name") or "",
                    "已删除",
                ]
            )
        ws2 = wb.create_sheet(tr('成绩'))
        ws2.append(["学号", "年份", "语文", "数学", "英语"])
        for r in scores:
            ws2.append(
                [
                    r["student_id"], r["year"],
                    r["chinese_score"], r["math_score"], r["english_score"],
                ]
            )
        ws3 = wb.create_sheet(tr('教师'))
        ws3.append(["账号", "姓名", "角色", "工号", "身份证号"])
        for t in teachers:
            ws3.append(
                [
                    t["username"], t.get("real_name") or "",
                    tr('管理员') if t.get("role") == "admin" else tr('教师'),
                    t.get("employee_id") or "", t.get("id_card") or "",
                ]
            )
        ws4 = wb.create_sheet(tr('日志'))
        ws4.append(["时间", "操作人", "角色", "操作", "对象", "详情"])
        for r in logs:
            ws4.append(
                [
                    r["created_at"], r["operator"], r["role"],
                    r["action"], r.get("target") or "", r.get("detail") or "",
                ]
            )
        wb.save(path)
        wb.close()
